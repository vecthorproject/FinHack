import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import zipfile
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from report_corp import genera_report_word
from report_breve_corp import genera_presentazione_ppt

# ==========================================
# IMPOSTAZIONI PAGINA WEB E GRAFICA (CSS)
# ==========================================
st.set_page_config(page_title="FinHack", page_icon="💹", layout="wide")

# --- INIEZIONE CSS PER TEMA SCURO PREMIUM ---
st.markdown("""
<style>
    /* Sfondo principale: Gradiente diagonale elegante (Midnight Blue to Deep Black) */
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #0B101E 0%, #05070A 100%);
        background-attachment: fixed;
    }
    
    /* Colore testo base globale */
    .stMarkdown, p, h1, h2, h3, span {
        color: #F8FAFC !important;
    }

    /* TITOLO PRINCIPALE: Dimensioni corrette e nessun accavallamento */
    .main-title {
        font-size: 3.0rem !important; /* Aumentato da 4.5 a 6.0 */
        color: #FFFFFF !important; 
        text-align: center;
        font-weight: 900;
        margin-top: 1rem;
        margin-bottom: 0px; 
        line-height: 1.1; /* Lo tiene compatto anche se è gigante */
        font-family: 'Inter', 'Helvetica Neue', Arial, sans-serif;
        text-shadow: 0px 4px 20px rgba(56, 189, 248, 0.15); 
    }
    
    /* SOTTOTITOLO: Distanziato e pulito */
    .sub-title {
        text-align: center;
        color: #94A3B8 !important; 
        font-size: 1.5rem;
        margin-top: 8px;
        margin-bottom: 3.5rem; /* Tanto respiro prima della linea divisoria */
        font-weight: 300;
        letter-spacing: 0.5px;
    }
    
    /* Box Messaggi (Effetto Vetro / Glassmorphism) */
    div.stAlert > div {
        border-radius: 12px;
        font-size: 1.1rem;
        background: rgba(30, 41, 59, 0.3) !important; /* Semi-trasparente */
        backdrop-filter: blur(10px); /* Sfoca quello che c'è dietro */
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.2);
    }
    
    /* Barra di caricamento file (Effetto Vetro Hover) */
    [data-testid="stFileUploadDropzone"] {
        border: 2px dashed #38BDF8 !important;
        border-radius: 16px;
        background: rgba(15, 23, 42, 0.4) !important;
        backdrop-filter: blur(5px);
        padding: 3rem !important; 
        transition: all 0.3s ease; /* Animazione morbida */
    }
    [data-testid="stFileUploadDropzone"]:hover {
        border-color: #7DD3FC !important;
        background: rgba(15, 23, 42, 0.7) !important;
    }
    [data-testid="stFileUploadDropzone"] * {
        color: #F8FAFC !important;
    }
    
    /* NUMERI DASHBOARD (Ridotti per farci stare le descrizioni) */
    [data-testid="stMetricValue"] {
        font-size: 2.5rem !important; 
        font-weight: 800 !important;
        color: #38BDF8 !important;
    }
    /* ETICHETTE DASHBOARD */
    [data-testid="stMetricLabel"] p {
        font-size: 1.1rem !important;
        font-weight: 600 !important;
        color: #CBD5E1 !important;
    }
    /* DETTAGLIO DELTA (I piccoli numerini sotto) */
    [data-testid="stMetricDelta"] {
        font-size: 0.9rem !important;
    }
</style>
""", unsafe_allow_html=True)

# Titoli visivi (Uso <div> al posto di <p> per evitare i margini imposti da Streamlit)
st.markdown('<p class="main-title">📊 FinHack ☠️</p>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Generazione avanzata e automatizzata dei Report Economico-Finanziari</div>', unsafe_allow_html=True)
st.divider()

st.info("💡 **ISTRUZIONI:** Carica un export in formato `.xlsx`. Assicurati di aver usato i filtri corretti e il formato **LISTA UNIVERSAL**.")

# ==========================================
# 1. FUNZIONI DEI CAPITOLI (Moduli)
# ==========================================

def elabora_capitolo_1(df_filtered, azienda_target):
    # ==========================================
    # GANCI INIZIALI
    # ==========================================
    df = df_filtered.copy()

    # --- LOGICA PANDAS DEL CAPITOLO 1 ---
    mappatura_fg = {
        'S.R.L. - SRL': 'Società a Responsabilità Limitata (S.r.l)',
        'Società a responsabilità limitata a capitale ridotto': 'Società a Responsabilità Limitata (S.r.l)',
        'S.R.L. semplificata': 'Società a Responsabilità Limitata (S.r.l)',
        'S.R.L. a socio unico - SRLU': 'Società a Responsabilità Limitata (S.r.l)',
        'Società europea - SE': 'Società per Azioni (S.p.A)',
        'Società per azioni - S.P.A. - SPA': 'Società per Azioni (S.p.A)',
        'S.A.P.A. - SAPA': 'Società per Azioni (S.p.A)',
        'S.P.A. a socio unico - SPA': 'Società per Azioni (S.p.A)'
    }

    colonna_fg = 'Forma giuridica nazionale' 
    df['Forma Giuridica Pulita'] = df[colonna_fg].str.replace(r'\s*\(Italia\)', '', regex=True).str.strip()
    df['Macro Forma Giuridica'] = df['Forma Giuridica Pulita'].map(mappatura_fg)
    df_cap1 = df.dropna(subset=['Forma Giuridica Pulita'])

    # 🟢 ESTRATTO TARGET: Individua la forma giuridica specifica e macro dell'azienda target
    df_target_check = df_cap1[df_cap1['Ragione socialeCaratteri latini'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    target_fg_pulita = df_target_check.iloc[0]['Forma Giuridica Pulita'] if not df_target_check.empty else None
    target_fg_macro = df_target_check.iloc[0]['Macro Forma Giuridica'] if not df_target_check.empty else None

    # FOGLIO "FG"
    fg_detail = df_cap1['Forma Giuridica Pulita'].value_counts().reset_index()
    fg_detail.columns = ['Forma giuridica', 'Conteggio di Ragione socialeCaratteri latini']
    fg_detail = fg_detail.sort_values('Forma giuridica').reset_index(drop=True)
    m_det = len(fg_detail)
    fg_detail.loc[m_det] = ['Totale complessivo', f'=SUM(B2:B{m_det+1})']

    fg_macro = df_cap1['Macro Forma Giuridica'].value_counts().reset_index()
    fg_macro.columns = ['Forma Giuridica', 'V.A.']
    m_mac = len(fg_macro)
    tot_row_excel = 4 + m_mac + 1 
    percents = [f'=F{5+i}/$F${tot_row_excel}*100' for i in range(m_mac)]
    fg_macro['%'] = percents
    fg_macro.loc[m_mac] = ['Totale', f'=SUM(F5:F{tot_row_excel-1})', f'=SUM(G5:G{tot_row_excel-1})']

    # FOGLIO "Liv.Agg. per FG"
    col_attivo = 'Totale Attivo migl EUR 2024'
    col_ricavi = 'Totale valore della produzione migl EUR 2024'

    fin_detail = df_cap1.groupby('Forma Giuridica Pulita')[[col_attivo, col_ricavi]].sum().reset_index()
    fin_detail.columns = ['Etichette di riga', 'Somma di Totale Attivo migl EUR 2024', 'Somma di Totale valore della produzione migl EUR 2024']
    fin_detail['Somma di Totale Attivo migl EUR 2024'] = fin_detail['Somma di Totale Attivo migl EUR 2024'].round(2)
    fin_detail['Somma di Totale valore della produzione migl EUR 2024'] = fin_detail['Somma di Totale valore della produzione migl EUR 2024'].round(2)
    fin_detail = fin_detail.sort_values('Etichette di riga').reset_index(drop=True)
    m_fin_det = len(fin_detail)
    fin_detail.loc[m_fin_det] = ['Totale complessivo', f'=SUM(B2:B{m_fin_det+1})', f'=SUM(C2:C{m_fin_det+1})']

    grouped_fin = df_cap1.groupby('Macro Forma Giuridica')[[col_attivo, col_ricavi]].sum().T
    grouped_fin.index = ['Totale Attivo', 'Totale Ricavi']

    fin_macro_temp = pd.DataFrame(index=grouped_fin.index)
    fin_macro_temp['S.r.l V.A.'] = grouped_fin.get('Società a Responsabilità Limitata (S.r.l)', 0).round(2)
    fin_macro_temp['S.p.A. V.A.'] = grouped_fin.get('Società per Azioni (S.p.A)', 0).round(2)
    fin_macro_temp.reset_index(inplace=True)
    fin_macro_temp.rename(columns={'index': 'Variabile'}, inplace=True)
    fin_macro_temp = fin_macro_temp.iloc[::-1].reset_index(drop=True)

    fin_macro = pd.DataFrame()
    fin_macro['Variabile'] = fin_macro_temp['Variabile']
    fin_macro['S.r.l V.A.'] = fin_macro_temp['S.r.l V.A.']
    
    srl_perc, spa_perc, tot_va, tot_perc = [], [], [], []
    for i in range(len(fin_macro_temp)):
        r = 5 + i 
        tot_va.append(f'=F{r}+H{r}')
        srl_perc.append(f'=IF(J{r}=0,0,F{r}/J{r}*100)')
        spa_perc.append(f'=IF(J{r}=0,0,H{r}/J{r}*100)')
        tot_perc.append(f'=G{r}+I{r}')

    fin_macro['S.r.l %'] = srl_perc
    fin_macro['S.p.A. V.A.'] = fin_macro_temp['S.p.A. V.A.']
    fin_macro['S.p.A. %'] = spa_perc
    fin_macro['Totale V.A.'] = tot_va
    fin_macro['Totale %'] = tot_perc


    # ==========================================
    # GANCIO DI MEZZO (Scrive in RAM invece che su disco)
    # ==========================================
    # Estrazione riga isolata dell'azienda target
    df_target_cap1 = df_cap1[df_cap1['Ragione socialeCaratteri latini'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    if not df_target_cap1.empty:
        forma_p = df_target_cap1.iloc[0]['Forma Giuridica Pulita']
        macro_p = df_target_cap1.iloc[0]['Macro Forma Giuridica']
        info_target_data = pd.DataFrame({
            'Metrica Target': ['Ragione Sociale', 'Forma Giuridica Specifica', 'Macro Categoria Appartenenza'],
            'Valore': [azienda_target, forma_p, macro_p]
        })
    else:
        info_target_data = pd.DataFrame({'Nota': ['Azienda Target non trovata nel campione corrente']})

    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        fg_detail.to_excel(writer, sheet_name='FG', index=False, startcol=0, startrow=0)
        fg_macro.to_excel(writer, sheet_name='FG', index=False, startcol=4, startrow=3)
        fin_detail.to_excel(writer, sheet_name='Liv.Agg. per FG', index=False, startcol=0, startrow=0)
        fin_macro.to_excel(writer, sheet_name='Liv.Agg. per FG', index=False, startcol=4, startrow=3)
        # Scrittura del foglio di isolamento dedicato
        info_target_data.to_excel(writer, sheet_name='Target_Forma_Giuridica', index=False)

    # ==========================================
    # GANCIO DI FORMATTAZIONE
    # ==========================================
    output_buffer.seek(0)
    wb = load_workbook(output_buffer)

    header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def format_table(worksheet, start_row, start_col, dataframe, is_count=False, has_total_row=True, highlight_value=None):
        end_row = start_row + len(dataframe)
        end_col = start_col + len(dataframe.columns) - 1

        # Stili di evidenziazione gialla per la riga ministeriale o societaria corrispondente
        highlight_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        highlight_font = Font(bold=True, color='000000')

        for row in range(start_row, end_row + 1):
            # Controlla se la prima cella descrittiva della riga combacia con il nostro target
            is_highlight = False
            if row > start_row and highlight_value is not None:
                cell_descr = str(worksheet.cell(row=row, column=start_col).value).strip()
                if cell_descr.lower() == str(highlight_value).lower().strip():
                    is_highlight = True

            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if row == start_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.border = thin_border
                    if is_highlight and not (has_total_row and row == end_row):
                        # Colora l'intera riga di giallo se appartiene alla categoria dell'azienda target
                        cell.fill = highlight_fill
                        cell.font = highlight_font
                    else:
                        if row % 2 != 0:
                            cell.fill = alt_row_fill
                        if has_total_row and row == end_row:
                            cell.font = total_font
                            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

                if row > start_row and col > start_col:
                    header_val = str(worksheet.cell(row=start_row, column=col).value)
                    if "%" in header_val:
                        cell.number_format = '0.00'
                    elif is_count:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

    ws_fg = wb['FG']
    # Evidenzia la forma giuridica specifica (es. Società per azioni) e quella macro
    format_table(ws_fg, 1, 1, fg_detail, is_count=True, has_total_row=True, highlight_value=target_fg_pulita)
    format_table(ws_fg, 4, 5, fg_macro, is_count=True, has_total_row=True, highlight_value=target_fg_macro)

    ws_fin = wb['Liv.Agg. per FG']
    format_table(ws_fin, 1, 1, fin_detail, is_count=False, has_total_row=True, highlight_value=target_fg_pulita)
    format_table(ws_fin, 4, 5, fin_macro, is_count=False, has_total_row=False) # Mantiene standard

    ws_target_1 = wb['Target_Forma_Giuridica']
    format_table(ws_target_1, 1, 1, info_target_data, is_count=False, has_total_row=False)

    for ws in [ws_fg, ws_fin, ws_target_1]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min((max_length + 2), 50)


    # ==========================================
    # GANCI FINALI
    # ==========================================
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output



def elabora_capitolo_2(df_filtered, azienda_target):
    import io
    import pandas as pd

    # ==========================================
    # GANCI INIZIALI
    # ==========================================
    df_base = df_filtered.copy()

    # --- LOGICA PANDAS ---
    cols = [
        'Ragione socialeCaratteri latini', 'NUTS2', 'Totale Attivo migl EUR 2024', 
        'Totale valore della produzione migl EUR 2024', 'Numero dipendenti 2024'
    ]
    # Filtriamo solo le colonne che esistono realmente nel file
    cols_to_use = [c for c in cols if c in df_base.columns]
    df_base = df_base[cols_to_use].copy()
    
    df_base.rename(columns={
        'NUTS2': 'Regione', 
        'Totale valore della produzione migl EUR 2024': 'Totale Ricavi migl EUR 2024',
        'Ragione socialeCaratteri latini': 'Ragione Sociale'
    }, inplace=True)

    def pulisci_regione(nome_regione):
        if pd.isna(nome_regione): return 'Altro'
        if ' - ' in str(nome_regione): return str(nome_regione).split(' - ', 1)[1]
        return str(nome_regione)
        
    if 'Regione' in df_base.columns:
        df_base['Nome Regione'] = df_base['Regione'].apply(pulisci_regione)
    else:
        df_base['Nome Regione'] = 'Altro'

    for col in ['Totale Attivo migl EUR 2024', 'Totale Ricavi migl EUR 2024']:
        if col in df_base.columns:
            df_base[col] = pd.to_numeric(df_base[col], errors='coerce')
            
    if 'Numero dipendenti 2024' in df_base.columns:
        df_base['Numero dipendenti 2024'] = pd.to_numeric(df_base['Numero dipendenti 2024'], errors='coerce')

    def get_macro(nuts2):
        if pd.isna(nuts2): return 'Altro'
        code = str(nuts2)[:3]
        if code == 'ITC': return 'Nord Ovest'
        elif code == 'ITH': return 'Nord Est'
        elif code == 'ITI': return 'Centro'
        elif code in ['ITF', 'ITG']: return 'Sud e Isole'
        else: return 'Altro'

    if 'Regione' in df_base.columns:
        df_base['Macroregione'] = df_base['Regione'].apply(get_macro)
    else:
         df_base['Macroregione'] = 'Altro'

    pivot_reg = df_base.groupby(['Macroregione', 'Nome Regione']).agg({
        'Ragione Sociale': 'count',
        'Totale Ricavi migl EUR 2024': 'sum',
        'Totale Attivo migl EUR 2024': 'sum',
        'Numero dipendenti 2024': 'sum'
    }).rename(columns={'Ragione Sociale': 'Imprese'})

    pivot_reg['Totale Ricavi migl EUR 2024'] = pivot_reg['Totale Ricavi migl EUR 2024'].round(2)
    pivot_reg['Totale Attivo migl EUR 2024'] = pivot_reg['Totale Attivo migl EUR 2024'].round(2)
    pivot_reg['Numero dipendenti 2024'] = pivot_reg['Numero dipendenti 2024'].fillna(0).astype(int)
    pivot_reg['Imprese'] = pivot_reg['Imprese'].fillna(0).astype(int)

    # 🟢 ESTRATTO TARGET: Trova la regione (NUTS2) e la macroregione dell'azienda bersaglio
    df_az_geo_check = df_base[df_base['Ragione Sociale'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    target_regione_nome = df_az_geo_check.iloc[0]['Nome Regione'] if not df_az_geo_check.empty else None
    target_macro_nome = df_az_geo_check.iloc[0]['Macroregione'] if not df_az_geo_check.empty else None

    # ==========================================
    # GANCIO DI MEZZO E FORMATTAZIONE CON XLSXWRITER
    # ==========================================
    output_buffer = io.BytesIO()
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book
    worksheet = workbook.add_worksheet('Dati Aggregate')

    # 🟢 STILI HIGHLIGHT: Creazione dei formati di riga gialla per xlsxwriter
    fmt_hl_text = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'bold': True})
    fmt_hl_int  = workbook.add_format({'num_format': '#,##0', 'border': 1, 'bg_color': '#FFF2CC', 'bold': True})
    fmt_hl_dec  = workbook.add_format({'num_format': '#,##0.00', 'border': 1, 'bg_color': '#FFF2CC', 'bold': True})
    fmt_hl_perc = workbook.add_format({'num_format': '0.00%', 'border': 1, 'bg_color': '#FFF2CC', 'bold': True})

    # Palette Stili
    format_header_blue = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
    format_subheader = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#D9D9D9', 'border': 1})
    format_regione = workbook.add_format({'border': 1})
    format_macro_bold = workbook.add_format({'bold': True, 'fg_color': '#F2F2F2', 'border': 1})
    format_italia = workbook.add_format({'bold': True, 'fg_color': '#D9E1F2', 'top': 1, 'bottom': 6})
    
    f_int = workbook.add_format({'num_format': '#,##0', 'border': 1})
    f_dec = workbook.add_format({'num_format': '#,##0.00', 'border': 1})
    f_perc = workbook.add_format({'num_format': '0.00%', 'border': 1})
    
    f_macro_int = workbook.add_format({'bold': True, 'num_format': '#,##0', 'fg_color': '#F2F2F2', 'border': 1})
    f_macro_dec = workbook.add_format({'bold': True, 'num_format': '#,##0.00', 'fg_color': '#F2F2F2', 'border': 1})
    f_macro_perc = workbook.add_format({'bold': True, 'num_format': '0.00%', 'fg_color': '#F2F2F2', 'border': 1})

    f_ita_int = workbook.add_format({'bold': True, 'num_format': '#,##0', 'fg_color': '#D9E1F2', 'top': 1, 'bottom': 6})
    f_ita_dec = workbook.add_format({'bold': True, 'num_format': '#,##0.00', 'fg_color': '#D9E1F2', 'top': 1, 'bottom': 6})
    f_ita_perc = workbook.add_format({'bold': True, 'num_format': '0.00%', 'fg_color': '#D9E1F2', 'top': 1, 'bottom': 6})

    # Intestazioni unite
    worksheet.write(0, 0, 'Regioni', format_header_blue)
    worksheet.merge_range(0, 1, 0, 2, 'Numero Imprese', format_header_blue)
    worksheet.merge_range(0, 3, 0, 4, 'Somma di Ricavi - migl EUR', format_header_blue)
    worksheet.merge_range(0, 5, 0, 6, 'Somma di Totale Attivo - migl EUR', format_header_blue)
    worksheet.merge_range(0, 7, 0, 8, 'Numero Dipendenti', format_header_blue)
    
    subheaders = ['', 'V.A.', '%', 'V.A.', '%', 'V.A.', '%', 'V.A.', '%']
    for col_num, sh in enumerate(subheaders):
        if col_num > 0: worksheet.write(1, col_num, sh, format_subheader)
        else: worksheet.write(1, col_num, sh)

    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:I', 18)

    macros_order = ['Nord Ovest', 'Nord Est', 'Centro', 'Sud e Isole']
    macros_present = [m for m in macros_order if m in pivot_reg.index.get_level_values(0)]
    
    num_regions = len(pivot_reg)
    num_macros = len(macros_present)
    riga_italia_idx = 2 + num_regions + num_macros
    riga_italia_excel = riga_italia_idx + 1

    current_idx = 2
    macro_tot_rows = []

    for macro in macros_present:
        df_macro = pivot_reg.loc[macro]
        start_reg_idx = current_idx

        # Ciclo delle singole Regioni
        for reg, row_data in df_macro.iterrows():
            # Controlla se la regione corrente è quella dell'azienda target
            is_reg_target = (target_regione_nome is not None and str(reg).lower().strip() == str(target_regione_nome).lower().strip())

            # Sceglie al volo il formato (Giallo se target, standard altrimenti)
            f_r = fmt_hl_text if is_reg_target else format_regione
            f_i = fmt_hl_int if is_reg_target else f_int
            f_d = fmt_hl_dec if is_reg_target else f_dec
            f_p = fmt_hl_perc if is_reg_target else f_perc

            worksheet.write(current_idx, 0, reg, f_r)
            worksheet.write(current_idx, 1, row_data['Imprese'], f_i)
            worksheet.write(current_idx, 3, row_data['Totale Ricavi migl EUR 2024'], f_d)
            worksheet.write(current_idx, 5, row_data['Totale Attivo migl EUR 2024'], f_d)
            worksheet.write(current_idx, 7, row_data['Numero dipendenti 2024'], f_i)
            for c, v_col in zip([2, 4, 6, 8], ['B', 'D', 'F', 'H']):
                worksheet.write_formula(current_idx, c, f"={v_col}{current_idx+1}/{v_col}${riga_italia_excel}", f_p)
            current_idx += 1

        end_reg_idx = current_idx - 1

        # Totale della Macroregione
        is_macro_target = (target_macro_nome is not None and str(macro).lower().strip() == str(target_macro_nome).lower().strip())

        # Formati dedicati con sfondo giallo per il sub-totale della macroregione dell'azienda target
        f_m_b = workbook.add_format({'bold': True, 'bg_color': '#FFF2CC', 'border': 1}) if is_macro_target else format_macro_bold
        f_m_i = workbook.add_format({'bold': True, 'num_format': '#,##0', 'bg_color': '#FFF2CC', 'border': 1}) if is_macro_target else f_macro_int
        f_m_d = workbook.add_format({'bold': True, 'num_format': '#,##0.00', 'bg_color': '#FFF2CC', 'border': 1}) if is_macro_target else f_macro_dec
        f_m_p = workbook.add_format({'bold': True, 'num_format': '0.00%', 'bg_color': '#FFF2CC', 'border': 1}) if is_macro_target else f_macro_perc

        worksheet.write(current_idx, 0, macro, f_m_b)
        for c, v_col in zip([1, 3, 5, 7], ['B', 'D', 'F', 'H']):
            worksheet_formula = f"=SUM({v_col}{start_reg_idx+1}:{v_col}{end_reg_idx+1})"
            worksheet.write_formula(current_idx, c, worksheet_formula, f_m_i if c in [1,7] else f_m_d)
        for c, v_col in zip([2, 4, 6, 8], ['B', 'D', 'F', 'H']):
            worksheet.write_formula(current_idx, c, f"={v_col}{current_idx+1}/{v_col}${riga_italia_excel}", f_m_p)
        macro_tot_rows.append(current_idx)
        current_idx += 1

    worksheet.write(current_idx, 0, "Italia", format_italia)
    for c, v_col in zip([1, 3, 5, 7], ['B', 'D', 'F', 'H']):
        if macro_tot_rows:
            sum_str = "+".join([f"{v_col}{r+1}" for r in macro_tot_rows])
            worksheet.write_formula(current_idx, c, f"={sum_str}", f_ita_int if c in [1,7] else f_ita_dec)
    for c, v_col in zip([2, 4, 6, 8], ['C', 'E', 'G', 'I']):
        if macro_tot_rows:
            sum_str = "+".join([f"{v_col}{r+1}" for r in macro_tot_rows])
            worksheet.write_formula(current_idx, c, f"={sum_str}", f_ita_perc)

    col_macro_start = 11 
    worksheet.write(1, col_macro_start, 'Macroregione', format_header_blue)
    worksheet.write(1, col_macro_start+1, 'Imprese %', format_header_blue)
    worksheet.write(1, col_macro_start+2, 'Ricavi %', format_header_blue)
    worksheet.write(1, col_macro_start+3, 'Attivo %', format_header_blue)
    worksheet.write(1, col_macro_start+4, 'Dipendenti %', format_header_blue)

    for i, (macro, m_row) in enumerate(zip(macros_present, macro_tot_rows)):
        r = 2 + i
        worksheet.write(r, col_macro_start, macro)
        worksheet.write_formula(r, col_macro_start+1, f"=C{m_row+1}", f_perc)
        worksheet.write_formula(r, col_macro_start+2, f"=E{m_row+1}", f_perc)
        worksheet.write_formula(r, col_macro_start+3, f"=G{m_row+1}", f_perc)
        worksheet.write_formula(r, col_macro_start+4, f"=I{m_row+1}", f_perc)

    col_reg_start = 17 
    worksheet.write(1, col_reg_start, 'Regione', format_header_blue)
    worksheet.write(1, col_reg_start+1, 'Numero Imprese', format_header_blue)
    worksheet.write(1, col_reg_start+2, 'Tot. Ricavi - migl EUR', format_header_blue)
    worksheet.write(1, col_reg_start+3, 'Tot. Attivo - migl EUR', format_header_blue)
    worksheet.write(1, col_reg_start+4, 'Numero Dipendenti', format_header_blue)

    reg_idx = 2
    for r in range(2, riga_italia_idx):
        if r not in macro_tot_rows: 
            worksheet.write_formula(reg_idx, col_reg_start, f"=A{r+1}")
            worksheet.write_formula(reg_idx, col_reg_start+1, f"=B{r+1}", f_int)
            worksheet.write_formula(reg_idx, col_reg_start+2, f"=D{r+1}", f_dec)
            worksheet.write_formula(reg_idx, col_reg_start+3, f"=F{r+1}", f_dec)
            worksheet.write_formula(reg_idx, col_reg_start+4, f"=H{r+1}", f_int)
            reg_idx += 1

    worksheet.set_column('L:V', 18)

    def crea_torta(titolo, colonna_valori, pos_cella):
        chart = workbook.add_chart({'type': 'pie'})
        chart.add_series({
            'name':       titolo,
            'categories': ['Dati Aggregate', 2, col_macro_start, 2 + num_macros - 1, col_macro_start], 
            'values':     ['Dati Aggregate', 2, colonna_valori, 2 + num_macros - 1, colonna_valori],
            'data_labels': {'percentage': True, 'position': 'outside_end'}
        })
        chart.set_title({'name': titolo})
        chart.set_style(10)
        chart.set_size({'width': 400, 'height': 280}) 
        worksheet.insert_chart(pos_cella, chart) 

    def crea_istogramma(titolo, colonna_valori, pos_cella):
        chart = workbook.add_chart({'type': 'column'})
        chart.add_series({
            'name':       titolo,
            'categories': ['Dati Aggregate', 2, col_reg_start, reg_idx - 1, col_reg_start], 
            'values':     ['Dati Aggregate', 2, colonna_valori, reg_idx - 1, colonna_valori],
        })
        chart.set_title({'name': titolo})
        chart.set_legend({'none': True})
        chart.set_style(10)
        chart.set_size({'width': 550, 'height': 280}) 
        worksheet.insert_chart(pos_cella, chart) 

    crea_torta('Ripartizione Imprese (%)', col_macro_start+1, 'L9')
    crea_torta('Ripartizione Tot. Ricavi (%)', col_macro_start+2, 'L24')
    crea_torta('Ripartizione Tot. Attivo (%)', col_macro_start+3, 'L39')
    crea_torta('Ripartizione Dipendenti (%)', col_macro_start+4, 'L54')

    col_chart_istogrammi = 'X'
    crea_istogramma('Imprese per Regione', col_reg_start+1, f'{col_chart_istogrammi}2')
    crea_istogramma('Tot. Ricavi per Regione (migl EUR)', col_reg_start+2, f'{col_chart_istogrammi}17')
    crea_istogramma('Tot. Attivo per Regione (migl EUR)', col_reg_start+3, f'{col_chart_istogrammi}32')
    crea_istogramma('Dipendenti per Regione', col_reg_start+4, f'{col_chart_istogrammi}47')

    ws_quartili = workbook.add_worksheet('Quartili')
    
    df_raw = df_base[['Totale Attivo migl EUR 2024', 'Totale Ricavi migl EUR 2024']].dropna()
    df_raw = df_raw.sort_values(by='Totale Ricavi migl EUR 2024', ascending=False)
    
    ws_quartili.write(0, 0, 'Totale Attivo migl EUR 2024', format_header_blue)
    ws_quartili.write(0, 1, 'Totale valore della produzione migl EUR 2024', format_header_blue)
    
    for r_idx, (_, row) in enumerate(df_raw.iterrows(), 1):
        ws_quartili.write(r_idx, 0, row['Totale Attivo migl EUR 2024'], f_dec)
        ws_quartili.write(r_idx, 1, row['Totale Ricavi migl EUR 2024'], f_dec)

    ws_quartili.set_column('A:B', 35)
    ultima_riga_dati = len(df_raw) + 1

    ws_quartili.write(0, 3, 'Variabile', format_header_blue)
    ws_quartili.write(0, 4, 'V.A.', format_header_blue)
    ws_quartili.write(1, 3, 'Totale Ricavi - migl EUR')
    ws_quartili.write_formula(1, 4, f"=SUM(B2:B{ultima_riga_dati})", f_dec)
    ws_quartili.write(2, 3, 'Totale Attivo - migl EUR')
    ws_quartili.write_formula(2, 4, f"=SUM(A2:A{ultima_riga_dati})", f_dec)

    ws_quartili.set_column('D:D', 22)
    ws_quartili.set_column('E:E', 20)

    ws_quartili.write(0, 6, 'Quartile', format_header_blue)
    ws_quartili.write(0, 7, 'Totale Attivo - migl EUR', format_header_blue)
    ws_quartili.write(0, 8, 'Totale Ricavi - migl EUR', format_header_blue)

    nomi_quartili = ['Minimo', '1°', '2°', '3°', '4°']
    for i, nome in enumerate(nomi_quartili):
        riga = i + 1
        ws_quartili.write(riga, 6, nome, format_subheader)
        ws_quartili.write_formula(riga, 7, f"=ROUND(QUARTILE(A2:A{ultima_riga_dati}, {i}), 2)", f_dec)
        ws_quartili.write_formula(riga, 8, f"=ROUND(QUARTILE(B2:B{ultima_riga_dati}, {i}), 2)", f_dec)

    start_r = 8
    ws_quartili.write(start_r, 6, 'Quartile', format_header_blue)
    ws_quartili.write(start_r, 7, 'Totale Attivo - migl EUR', format_header_blue)
    ws_quartili.write(start_r, 8, 'Totale Ricavi - migl EUR', format_header_blue)

    intervalli_nomi = ['1°', '2°', '3°', '4°']
    for i in range(4):
        par_chiusa = ')' if i < 3 else ']'
        riga_min = i + 2 
        riga_max = i + 3 
        formula_attivo = f'="[" & TEXT(H{riga_min}, "#.##0,00") & " - " & TEXT(H{riga_max}, "#.##0,00") & "{par_chiusa}"'
        formula_ricavi = f'="[" & TEXT(I{riga_min}, "#.##0,00") & " - " & TEXT(I{riga_max}, "#.##0,00") & "{par_chiusa}"'
        
        ws_quartili.write(start_r + 1 + i, 6, intervalli_nomi[i], format_subheader)
        ws_quartili.write_formula(start_r + 1 + i, 7, formula_attivo)
        ws_quartili.write_formula(start_r + 1 + i, 8, formula_ricavi)

    ws_quartili.set_column('G:I', 40)

    # ==========================================
    # GANCI FINALI
    # ==========================================
    # Isolamento dati geografici e dimensionali dell'azienda target
    df_az_geo = df_base[df_base['Ragione Sociale'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    ws_target_geo = workbook.add_worksheet('Target_Posizionamento_Geo')

    ws_target_geo.set_column('A:A', 40)
    ws_target_geo.set_column('B:B', 35)

    ws_target_geo.write(0, 0, 'Anagrafica / Geometria Territorial', format_header_blue)
    ws_target_geo.write(0, 1, 'Dettaglio Azienda Target', format_header_blue)

    if not df_az_geo.empty:
        riga_g = df_az_geo.iloc[0]
        voci_geo = [
            ('Ragione Sociale', riga_g['Ragione Sociale'], format_regione),
            ('Macroregione Appartenenza', riga_g['Macroregione'], format_regione),
            ('Regione Specifica (NUTS2)', riga_g['Nome Regione'], format_regione),
            ('Totale Ricavi - migl EUR 2024', riga_g['Totale Ricavi migl EUR 2024'], f_dec),
            ('Totale Attivo - migl EUR 2024', riga_g['Totale Attivo migl EUR 2024'], f_dec),
            ('Numero Dipendenti 2024', riga_g['Numero dipendenti 2024'], f_int)
        ]
        
        for idx, (voce, valore, formato_cella) in enumerate(voci_geo, 1):
            ws_target_geo.write(idx, 0, voce, format_macro_bold)
            
            # --- FIX ANTI-CRASH PER I DATI MANCANTI ---
            if pd.isna(valore) or str(valore).strip() == "":
                ws_target_geo.write(idx, 1, "n.d.", formato_cella)
            else:
                ws_target_geo.write(idx, 1, valore, formato_cella)
            
    else:
        ws_target_geo.write(1, 0, 'Azienda Target non trovata', format_regione)

    writer.close()
    output_buffer.seek(0)
    return output_buffer


def elabora_capitolo_3(df_filtered, azienda_target):
    import io
    import pandas as pd
    import numpy as np
    import xlsxwriter
    from xlsxwriter.utility import xl_rowcol_to_cell
    import re

    # Funzione interna che hai creato tu, adattata per girare qui dentro
    def costruisci_sezione_analisi(writer, workbook, df_raw, formati, keyword_ricerca, sheet_data, sheet_stats, chart_title, y_axis_name, rename_dict, azienda_target):
        # 1. Filtro Colonne per la sezione corrente
        base_cols = [c for c in df_raw.columns if 'ragione' in str(c).lower() or 'bvd' in str(c).lower() or 'nuts2' in str(c).lower() or 'nuts3' in str(c).lower()]
        metric_cols = [c for c in df_raw.columns if keyword_ricerca in str(c).lower()]
        
        if not metric_cols:
            return

        df = df_raw[base_cols + metric_cols].copy()
        df = df.replace(['n.d.', 'n.a.', 'n.s.', ''], np.nan)
        num_rows = len(df)

        col_ragione_idx = [idx for idx, c in enumerate(df.columns) if 'ragione' in str(c).lower()][0]
        col_ragione = df.columns[col_ragione_idx]

        # 2. Creazione Fogli
        df.to_excel(writer, sheet_name=sheet_data, index=False, startrow=0)
        worksheet_data = writer.sheets[sheet_data]
        worksheet_stats = workbook.add_worksheet(sheet_stats)

        # 3. Compilazione Foglio 1: Dati, Autofit ed EVIDENZIAZIONE AZIENDA TARGET
        fmt_target_text = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'left', 'valign': 'vcenter', 'bold': True})
        fmt_target_num = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'bold': True})

        for i, col in enumerate(df.columns):
            worksheet_data.write(0, i, col, formati['header'])
            col_data = df[col].dropna()
            max_len = len(str(col)) + 2 if col_data.empty else max(col_data.astype(str).map(len).max(), len(str(col))) + 2
            worksheet_data.set_column(i, i, min(max_len, 45))

            for row in range(1, num_rows + 1):
                val = df.iat[row-1, i]
                is_target = azienda_target.lower().strip() in str(df.iat[row-1, col_ragione_idx]).lower().strip()

                f_text = fmt_target_text if is_target else formati['data_text']
                f_num = fmt_target_num if is_target else formati['data_num']

                if pd.isna(val):
                    worksheet_data.write(row, i, "", f_text)
                elif isinstance(val, (int, float)):
                    worksheet_data.write_number(row, i, val, f_num)
                else:
                    worksheet_data.write_string(row, i, str(val), f_text)

        # 4. Preparazione Metriche per il Foglio 2
        numeric_cols_idx = [df.columns.get_loc(c) for c in metric_cols]
        metrics_dict = {} 
        
        for col_idx in numeric_cols_idx:
            col_name = df.columns[col_idx]
            anno_match = re.search(r'\d{4}', col_name)
            if anno_match:
                anno = int(anno_match.group(0))
                metric_base = re.sub(r'\d{4}', '', col_name).replace('(*) %', '').replace('migl EUR', '').replace('  ', ' ').strip()
                for k, v in rename_dict.items():
                    if k.lower() in metric_base.lower():
                        metric_base = v
                
                if metric_base not in metrics_dict: metrics_dict[metric_base] = {}
                metrics_dict[metric_base][anno] = col_idx

        all_years = sorted(list(set(anno for m in metrics_dict for anno in metrics_dict[m].keys())))
        metrics_list = list(metrics_dict.keys())
        
        # --- NUOVA REGOLA: FORZA L'ORDINE SU EBITDA -> EBIT -> PROFITTO ---
        ordine_desiderato = ['ebitda', 'ebit', 'profit']
        metrics_list = sorted(metrics_list, key=lambda x: next((i for i, k in enumerate(ordine_desiderato) if k in x.lower()), 99))
        
        col_left = 0
        col_mid = len(all_years) + 2
        col_right = col_mid + (len(all_years)*2 - 1) + 2

        worksheet_stats.set_column(col_left, col_left, 22)
        worksheet_stats.set_column(col_left + 1, col_left + len(all_years), 12)
        worksheet_stats.set_column(col_mid, col_mid, 22)
        worksheet_stats.set_column(col_mid + 1, col_mid + (len(all_years)*2), 10)
        worksheet_stats.set_column(col_right, col_right, 22)
        worksheet_stats.set_column(col_right + 1, col_right + len(all_years), 12)

        median_cells = {}
        current_row = 1
        stats_formulas = [('Media', 'AVERAGE'), ('Mediana', 'MEDIAN'), ('Asimmetria', 'SKEW'), ('Curtosi', 'KURT'), ('Deviazione standard', 'STDEV')]

        # 5. Costruzione Blocco Sinistro (Stats) e Centrale (YoY)
        for metric in metrics_list:
            median_cells[metric] = {}
            
            # --- BLOCCO SINISTRO ---
            worksheet_stats.write(current_row, col_left, "", formati['header'])
            for i, year in enumerate(all_years):
                worksheet_stats.write(current_row, col_left + 1 + i, year, formati['header'])
                
            for s_idx, (stat_name, func) in enumerate(stats_formulas):
                r = current_row + 1 + s_idx
                worksheet_stats.write(r, col_left, stat_name, formati['label'])
                for i, year in enumerate(all_years):
                    if year in metrics_dict[metric]:
                        data_col_idx = metrics_dict[metric][year]
                        col_letter = xlsxwriter.utility.xl_col_to_name(data_col_idx)
                        data_range = f"'{sheet_data}'!{col_letter}2:{col_letter}{num_rows+1}"
                        formula = f"={func}({data_range})"
                        worksheet_stats.write_formula(r, col_left + 1 + i, formula, formati['data_num'])
                        if stat_name == 'Mediana':
                            median_cells[metric][year] = xl_rowcol_to_cell(r, col_left + 1 + i)
                    else:
                        worksheet_stats.write(r, col_left + 1 + i, "n.d.", formati['data_text'])

            # --- BLOCCO CENTRALE ---
            c_offset = 1
            for i, year in enumerate(all_years):
                if i == 0:
                    worksheet_stats.write(current_row, col_mid + c_offset, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    c_offset += 1
                else:
                    worksheet_stats.merge_range(current_row, col_mid + c_offset, current_row, col_mid + c_offset + 1, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset + 1, "Δ %", formati['header'])
                    c_offset += 2
                    
            r_data = current_row + 2
            worksheet_stats.write(r_data, col_mid, metric, formati['label'])
            c_offset = 1
            prev_target_cell = None
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    source_cell = median_cells[metric][year]
                    curr_target_cell = xl_rowcol_to_cell(r_data, col_mid + c_offset)
                    worksheet_stats.write_formula(r_data, col_mid + c_offset, f"={source_cell}", formati['data_num'])
                    
                    if i > 0:
                        if prev_target_cell:
                            delta_formula = f"=IFERROR(({curr_target_cell}-{prev_target_cell})/ABS({prev_target_cell}), 0)"
                            worksheet_stats.write_formula(r_data, col_mid + c_offset + 1, delta_formula, formati['pct'])
                        else:
                            worksheet_stats.write(r_data, col_mid + c_offset + 1, "n.d.", formati['data_text'])
                        prev_target_cell = curr_target_cell
                        c_offset += 2
                    else:
                        prev_target_cell = curr_target_cell
                        c_offset += 1

            current_row += 8 

        # 6. Costruzione Blocco Destro: CONFRONTO DIRETTO RIGHE SETTORE VS AZIENDA TARGET
        r_right = 1
        worksheet_stats.write(r_right, col_right, "Confronto Benchmark", formati['header'])
        for i, year in enumerate(all_years):
            worksheet_stats.write(r_right, col_right + 1 + i, year, formati['header'])

        df_azienda = df[df[col_ragione].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
        fmt_az_label = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'font_color': '#002060', 'border': 1, 'align': 'left'})
        fmt_az_num = workbook.add_format({'bg_color': '#F2F4F8', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'bold': True})

        row_cursor = r_right + 1
        for metric in metrics_list:
            # Riga 1: Settore
            worksheet_stats.write(row_cursor, col_right, f"{metric} (Settore)", formati['label'])
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    worksheet_stats.write_formula(row_cursor, col_right + 1 + i, f"={median_cells[metric][year]}", formati['data_num'])
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

            # Riga 2: Azienda Target
            worksheet_stats.write(row_cursor, col_right, f"{metric} ({azienda_target})", fmt_az_label)
            for i, year in enumerate(all_years):
                if year in metrics_dict[metric] and not df_azienda.empty:
                    val_az = df_azienda.iloc[0, metrics_dict[metric][year]]
                    if pd.isna(val_az) or val_az == "":
                        worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
                    else:
                        worksheet_stats.write_number(row_cursor, col_right + 1 + i, float(val_az), fmt_az_num)
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

        # ==========================================
        # 7. INSERIMENTO GRAFICI DINAMICI UNIFORMI (CAP 3, 4, 5)
        # ==========================================
        row_start_metrics = r_right + 1
        start_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + 1)
        end_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + len(all_years))
        last_year_col = end_col_letter 
        
        safe_chart_row = max(current_row, row_cursor)

        # 🟢 IMPOSTAZIONI ANTI-ACCAVALLAMENTO UNIVERSALI
        # Riduciamo i font per impedire che i numeri giganteschi sfondino i margini del grafico
        font_assi = {'size': 9}
        font_etichette = {'size': 8}
        
        # --- GRAFICO 1 (LINEE CUMULATIVE): Trend Settore ---
        chart_trend_settore = workbook.add_chart({'type': 'line'})
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            chart_trend_settore.add_series({
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}",
                'marker': {'type': 'circle'}
            })
        chart_trend_settore.set_title({'name': f'Andamento Storico di Settore ({chart_title.replace("Andamento Mediano ", "")})'})
        chart_trend_settore.set_legend({'position': 'bottom', 'font': font_assi})
        chart_trend_settore.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
        chart_trend_settore.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
        chart_trend_settore.set_size({'width': 650, 'height': 350})
        chart_trend_settore.set_style(11)
        worksheet_stats.insert_chart(safe_chart_row, col_left, chart_trend_settore)

        # --- GRAFICI MULTIPLI (COPPIE) ---
        chart_offset_y = safe_chart_row
        
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            r_azienda = r_settore + 1
            
            # 🟢 FIX ISTOGRAMMA: Se la metrica riguarda grandi volumi, abbandona la linea e crea un istogramma standard
            is_volumi = any(k in metric.lower() for k in ['ricavi', 'attivo', 'produzione'])
            tipo_storico = 'column' if is_volumi else 'line'

            # A) Grafico Storico (Linee o Colonne in base alla metrica)
            chart_storico = workbook.add_chart({'type': tipo_storico})
            
            # Setup posizioni etichette: se istogramma vanno "sopra il tetto" della barra, se linea le separiamo (Sopra/Sotto)
            pos_label_settore = 'outside_end' if is_volumi else 'below'
            pos_label_azienda = 'outside_end' if is_volumi else 'above'

            # 🟢 Disattiviamo i numeri fluttuanti: i valori precisi saranno nella Tabella in basso
            serie_settore_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}"
            }
            serie_azienda_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_azienda + 1}:${end_col_letter}${r_azienda + 1}"
            }

            # ... (codice marker rimane intatto) ...
            if not is_volumi:
                serie_settore_storico['marker'] = {'type': 'circle'}
                serie_azienda_storico['marker'] = {'type': 'square'}

            chart_storico.add_series(serie_azienda_storico)
            chart_storico.add_series(serie_settore_storico)
            chart_storico.set_title({'name': f'Benchmark Storico - {metric}'})

            # 🟢 Nessuna legenda necessaria: la tabella dati include già i nomi delle serie!
            chart_storico.set_legend({'none': True}) 

            # 🟢 ATTIVAZIONE MATRICE DATI (DATA TABLE)
            chart_storico.set_table({'show_keys': True})

            chart_storico.set_x_axis({'name_font': font_assi, 'num_font': font_assi})
            # 🟢 Togli la griglia orizzontale SOLO se il grafico storico è un istogramma (es. Ricavi)
            if is_volumi:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi, 'major_gridlines': {'visible': False}})
            else:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
            chart_storico.set_size({'width': 550, 'height': 350}) 
            chart_storico.set_style(11)
            
            # B) Grafico a Colonne / Istogramma (Azienda vs Settore - Solo anno 2024)
            chart_col_singolo = workbook.add_chart({'type': 'column'})
            
            serie_settore_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_settore + 1}:${last_year_col}${r_settore + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            serie_azienda_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_azienda + 1}:${last_year_col}${r_azienda + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            chart_col_singolo.add_series(serie_azienda_col)
            chart_col_singolo.add_series(serie_settore_col)
            chart_col_singolo.set_title({'name': f'Posizionamento {metric} (Anno 2024)'})
            chart_col_singolo.set_legend({'position': 'bottom', 'font': font_assi})
            chart_col_singolo.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
            chart_col_singolo.set_y_axis({'name_font': font_assi, 'num_font': font_assi, 'major_gridlines': {'visible': False}})
            chart_col_singolo.set_size({'width': 350, 'height': 350})
            chart_col_singolo.set_style(11)

            # Inseriamo i grafici distanziandoli a dovere sulla destra
            worksheet_stats.insert_chart(chart_offset_y, col_right + 1, chart_storico)
            worksheet_stats.insert_chart(chart_offset_y, col_right + 9, chart_col_singolo)
            
            # Incremento dello step verticale da 16 a 18 righe per distanziare nettamente una metrica dall'altra
            chart_offset_y += 20

        worksheet_stats.ignore_errors({'formula_differs': 'A1:Z500', 'number_stored_as_text': 'A1:Z500'})

    # ==========================================
    # LOGICA DI ESECUZIONE 
    # ==========================================
    df_raw = df_filtered.copy()
    output_buffer = io.BytesIO()
    
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book

    formati = {
        'header': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'}),
        'label': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'data_num': workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter'}),
        'data_text': workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'pct': workbook.add_format({'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})
    }

    costruisci_sezione_analisi(
        writer, workbook, df_raw, formati,
        keyword_ricerca='margin',
        sheet_data='3a. Dati Eq. Economico',
        sheet_stats='3b. Stat. Eq. Economico',
        chart_title='Andamento Mediano Margini',
        y_axis_name='Percentuale (%)',
        rename_dict={} ,
        azienda_target=azienda_target
    )

    costruisci_sezione_analisi(
        writer, workbook, df_raw, formati,
        keyword_ricerca='produzione', 
        sheet_data='4a. Dati Svil. Dimensionale',
        sheet_stats='4b. Stat. Svil. Dimensionale',
        chart_title='Andamento Mediano Ricavi',
        y_axis_name='Migliaia di Euro (€)',
        rename_dict={'produzione': 'Ricavi'} ,
        azienda_target=azienda_target
    )

    writer.close()
    output_buffer.seek(0)
    
    return output_buffer


def elabora_capitolo_4(df_filtered, azienda_target):
    import io
    import pandas as pd
    import numpy as np
    import xlsxwriter
    from xlsxwriter.utility import xl_rowcol_to_cell
    import re

    def costruisci_sezione_analisi(writer, workbook, df_raw, formati, keyword_ricerca, sheet_data, sheet_stats, chart_title, y_axis_name, rename_dict, azienda_target):
        # 1. Filtro Colonne
        base_cols = [c for c in df_raw.columns if 'ragione' in str(c).lower() or 'bvd' in str(c).lower() or 'nuts2' in str(c).lower() or 'nuts3' in str(c).lower()]
        
        if isinstance(keyword_ricerca, list):
            metric_cols = [c for c in df_raw.columns if any(kw.lower() in str(c).lower() for kw in keyword_ricerca)]
        else:
            metric_cols = [c for c in df_raw.columns if keyword_ricerca.lower() in str(c).lower()]

        if not metric_cols:
            return

        df = df_raw[base_cols + metric_cols].copy()
        df = df.replace(['n.d.', 'n.a.', 'n.s.', ''], np.nan)
        num_rows = len(df)

        col_ragione_idx = [idx for idx, c in enumerate(df.columns) if 'ragione' in str(c).lower()][0]
        col_ragione = df.columns[col_ragione_idx]

        # 2. Creazione Fogli
        df.to_excel(writer, sheet_name=sheet_data, index=False, startrow=0)
        worksheet_data = writer.sheets[sheet_data]
        worksheet_stats = workbook.add_worksheet(sheet_stats)

        # 3. Compilazione Foglio 1: Dati, Autofit ed EVIDENZIAZIONE AZIENDA TARGET
        fmt_target_text = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'left', 'valign': 'vcenter', 'bold': True})
        fmt_target_num = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'bold': True})

        for i, col in enumerate(df.columns):
            worksheet_data.write(0, i, col, formati['header'])
            col_data = df[col].dropna()
            max_len = len(str(col)) + 2 if col_data.empty else max(col_data.astype(str).map(len).max(), len(str(col))) + 2
            worksheet_data.set_column(i, i, min(max_len, 45))

            for row in range(1, num_rows + 1):
                val = df.iat[row-1, i]
                is_target = azienda_target.lower().strip() in str(df.iat[row-1, col_ragione_idx]).lower().strip()

                f_text = fmt_target_text if is_target else formati['data_text']
                f_num = fmt_target_num if is_target else formati['data_num']

                if pd.isna(val):
                    worksheet_data.write(row, i, "", f_text)
                elif isinstance(val, (int, float)):
                    worksheet_data.write_number(row, i, val, f_num)
                else:
                    worksheet_data.write_string(row, i, str(val), f_text)

        # 4. Preparazione Metriche per il Foglio 2
        numeric_cols_idx = [df.columns.get_loc(c) for c in metric_cols]
        metrics_dict = {}

        for col_idx in numeric_cols_idx:
            col_name = df.columns[col_idx]
            anno_match = re.search(r'\d{4}', col_name)
            if anno_match:
                anno = int(anno_match.group(0))
                metric_base = re.sub(r'\d{4}', '', col_name).replace('(*) %', '').replace('(*)', '').replace('migl EUR', '').replace(' ', ' ').strip()
                
                for k, v in rename_dict.items():
                    if k.lower() in metric_base.lower():
                        metric_base = v

                if metric_base not in metrics_dict: metrics_dict[metric_base] = {}
                metrics_dict[metric_base][anno] = col_idx

        all_years = sorted(list(set(anno for m in metrics_dict for anno in metrics_dict[m].keys())))
        metrics_list = list(metrics_dict.keys())

        col_left = 0
        col_mid = len(all_years) + 2
        col_right = col_mid + (len(all_years)*2 - 1) + 2

        worksheet_stats.set_column(col_left, col_left, 22)
        worksheet_stats.set_column(col_left + 1, col_left + len(all_years), 12)
        worksheet_stats.set_column(col_mid, col_mid, 22)
        worksheet_stats.set_column(col_mid + 1, col_mid + (len(all_years)*2), 10)
        worksheet_stats.set_column(col_right, col_right, 22)
        worksheet_stats.set_column(col_right + 1, col_right + len(all_years), 12)

        median_cells = {}
        current_row = 1
        stats_formulas = [('Media', 'AVERAGE'), ('Mediana', 'MEDIAN'), ('Asimmetria', 'SKEW'), ('Curtosi', 'KURT'), ('Deviazione standard', 'STDEV')]

        # 5. Costruzione Blocco Sinistro (Stats) e Centrale (YoY)
        for metric in metrics_list:
            median_cells[metric] = {}

            worksheet_stats.write(current_row, col_left, "", formati['header'])
            for i, year in enumerate(all_years):
                worksheet_stats.write(current_row, col_left + 1 + i, year, formati['header'])

            for s_idx, (stat_name, func) in enumerate(stats_formulas):
                r = current_row + 1 + s_idx
                worksheet_stats.write(r, col_left, stat_name, formati['label'])
                for i, year in enumerate(all_years):
                    if year in metrics_dict[metric]:
                        data_col_idx = metrics_dict[metric][year]
                        col_letter = xlsxwriter.utility.xl_col_to_name(data_col_idx)
                        data_range = f"'{sheet_data}'!{col_letter}2:{col_letter}{num_rows+1}"
                        formula = f"={func}({data_range})"
                        worksheet_stats.write_formula(r, col_left + 1 + i, formula, formati['data_num'])
                        if stat_name == 'Mediana':
                            median_cells[metric][year] = xl_rowcol_to_cell(r, col_left + 1 + i)
                    else:
                        worksheet_stats.write(r, col_left + 1 + i, "n.d.", formati['data_text'])

            c_offset = 1
            for i, year in enumerate(all_years):
                if i == 0:
                    worksheet_stats.write(current_row, col_mid + c_offset, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    c_offset += 1
                else:
                    worksheet_stats.merge_range(current_row, col_mid + c_offset, current_row, col_mid + c_offset + 1, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset + 1, "Δ %", formati['header'])
                    c_offset += 2

            r_data = current_row + 2
            worksheet_stats.write(r_data, col_mid, metric, formati['label'])
            c_offset = 1
            prev_target_cell = None
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    source_cell = median_cells[metric][year]
                    curr_target_cell = xl_rowcol_to_cell(r_data, col_mid + c_offset)
                    worksheet_stats.write_formula(r_data, col_mid + c_offset, f"={source_cell}", formati['data_num'])

                    if i > 0:
                        if prev_target_cell:
                            delta_formula = f"=IFERROR(({curr_target_cell}-{prev_target_cell})/ABS({prev_target_cell}), 0)"
                            worksheet_stats.write_formula(r_data, col_mid + c_offset + 1, delta_formula, formati['pct'])
                        else:
                            worksheet_stats.write(r_data, col_mid + c_offset + 1, "n.d.", formati['data_text'])
                        prev_target_cell = curr_target_cell
                        c_offset += 2
                    else:
                        prev_target_cell = curr_target_cell
                        c_offset += 1
                else:
                    prev_target_cell = curr_target_cell
                    c_offset += 1

            current_row += 8 

        # 6. Costruzione Blocco Destro: CONFRONTO DIRETTO RIGHE SETTORE VS AZIENDA TARGET
        r_right = 1
        worksheet_stats.write(r_right, col_right, "Confronto Benchmark", formati['header'])
        for i, year in enumerate(all_years):
            worksheet_stats.write(r_right, col_right + 1 + i, year, formati['header'])

        df_azienda = df[df[col_ragione].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
        fmt_az_label = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'font_color': '#002060', 'border': 1, 'align': 'left'})
        fmt_az_num = workbook.add_format({'bg_color': '#F2F4F8', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'bold': True})

        row_cursor = r_right + 1
        for metric in metrics_list:
            # Riga 1: Settore
            worksheet_stats.write(row_cursor, col_right, f"{metric} (Settore)", formati['label'])
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    worksheet_stats.write_formula(row_cursor, col_right + 1 + i, f"={median_cells[metric][year]}", formati['data_num'])
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

            # Riga 2: Azienda Target
            worksheet_stats.write(row_cursor, col_right, f"{metric} ({azienda_target})", fmt_az_label)
            for i, year in enumerate(all_years):
                if year in metrics_dict[metric] and not df_azienda.empty:
                    val_az = df_azienda.iloc[0, metrics_dict[metric][year]]
                    if pd.isna(val_az) or val_az == "":
                        worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
                    else:
                        worksheet_stats.write_number(row_cursor, col_right + 1 + i, float(val_az), fmt_az_num)
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

        # ==========================================
        # 7. INSERIMENTO GRAFICI DINAMICI UNIFORMI (CAP 3, 4, 5)
        # ==========================================
        row_start_metrics = r_right + 1
        start_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + 1)
        end_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + len(all_years))
        last_year_col = end_col_letter 
        
        safe_chart_row = max(current_row, row_cursor)

        # 🟢 IMPOSTAZIONI ANTI-ACCAVALLAMENTO UNIVERSALI
        # Riduciamo i font per impedire che i numeri giganteschi sfondino i margini del grafico
        font_assi = {'size': 9}
        font_etichette = {'size': 8}
        
        # --- GRAFICO 1 (LINEE CUMULATIVE): Trend Settore ---
        chart_trend_settore = workbook.add_chart({'type': 'line'})
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            chart_trend_settore.add_series({
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}",
                'marker': {'type': 'circle'}
            })
        chart_trend_settore.set_title({'name': f'Andamento Storico di Settore ({chart_title.replace("Andamento Mediano ", "")})'})
        chart_trend_settore.set_legend({'position': 'bottom', 'font': font_assi})
        chart_trend_settore.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
        chart_trend_settore.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
        chart_trend_settore.set_size({'width': 650, 'height': 350})
        chart_trend_settore.set_style(11)
        worksheet_stats.insert_chart(safe_chart_row, col_left, chart_trend_settore)

        # --- GRAFICI MULTIPLI (COPPIE) ---
        chart_offset_y = safe_chart_row
        
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            r_azienda = r_settore + 1
            
            # 🟢 FIX ISTOGRAMMA: Se la metrica riguarda grandi volumi, abbandona la linea e crea un istogramma standard
            is_volumi = any(k in metric.lower() for k in ['ricavi', 'attivo', 'produzione'])
            tipo_storico = 'column' if is_volumi else 'line'

            # A) Grafico Storico (Linee o Colonne in base alla metrica)
            chart_storico = workbook.add_chart({'type': tipo_storico})
            
            # Setup posizioni etichette: se istogramma vanno "sopra il tetto" della barra, se linea le separiamo (Sopra/Sotto)
            pos_label_settore = 'outside_end' if is_volumi else 'below'
            pos_label_azienda = 'outside_end' if is_volumi else 'above'

            # 🟢 Disattiviamo i numeri fluttuanti: i valori precisi saranno nella Tabella in basso
            serie_settore_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}"
            }
            serie_azienda_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_azienda + 1}:${end_col_letter}${r_azienda + 1}"
            }

            # ... (codice marker rimane intatto) ...
            if not is_volumi:
                serie_settore_storico['marker'] = {'type': 'circle'}
                serie_azienda_storico['marker'] = {'type': 'square'}

            chart_storico.add_series(serie_azienda_storico)
            chart_storico.add_series(serie_settore_storico)
            chart_storico.set_title({'name': f'Benchmark Storico - {metric}'})

            # 🟢 Nessuna legenda necessaria: la tabella dati include già i nomi delle serie!
            chart_storico.set_legend({'none': True}) 

            # 🟢 ATTIVAZIONE MATRICE DATI (DATA TABLE)
            chart_storico.set_table({'show_keys': True})

            chart_storico.set_x_axis({'name_font': font_assi, 'num_font': font_assi})
            # 🟢 Togli la griglia orizzontale SOLO se il grafico storico è un istogramma (es. Ricavi)
            if is_volumi:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi, 'major_gridlines': {'visible': False}})
            else:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
            chart_storico.set_size({'width': 550, 'height': 350}) 
            chart_storico.set_style(11)
            
            # B) Grafico a Colonne / Istogramma (Azienda vs Settore - Solo anno 2024)
            chart_col_singolo = workbook.add_chart({'type': 'column'})
            
            serie_settore_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_settore + 1}:${last_year_col}${r_settore + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            serie_azienda_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_azienda + 1}:${last_year_col}${r_azienda + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            chart_col_singolo.add_series(serie_azienda_col)
            chart_col_singolo.add_series(serie_settore_col)
            chart_col_singolo.set_title({'name': f'Posizionamento {metric} (Anno 2024)'})
            chart_col_singolo.set_legend({'position': 'bottom', 'font': font_assi})
            chart_col_singolo.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
            # 🟢 Rimuove la griglia orizzontale di sfondo dall'istogramma
            chart_col_singolo.set_y_axis({
                'name_font': font_assi, 
                'num_font': font_assi, 
                'major_gridlines': {'visible': False}
            })
            chart_col_singolo.set_size({'width': 350, 'height': 350})
            chart_col_singolo.set_style(11)

            # Inseriamo i grafici distanziandoli a dovere sulla destra
            worksheet_stats.insert_chart(chart_offset_y, col_right + 1, chart_storico)
            worksheet_stats.insert_chart(chart_offset_y, col_right + 9, chart_col_singolo)
            
            # Incremento dello step verticale da 16 a 18 righe per distanziare nettamente una metrica dall'altra
            chart_offset_y += 20 

        worksheet_stats.ignore_errors({'formula_differs': 'A1:Z500', 'number_stored_as_text': 'A1:Z500'})

    # ==========================================
    # LOGICA DI ESECUZIONE 
    # ==========================================
    df_raw = df_filtered.copy()
    output_buffer = io.BytesIO()
    
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book

    formati = {
        'header': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'}),
        'label': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'data_num': workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter'}),
        'data_text': workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'pct': workbook.add_format({'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})
    }

    costruisci_sezione_analisi(
        writer, workbook, df_raw, formati,
        keyword_ricerca=['struttura', 'gearing'], 
        sheet_data='5a. Dati Eq. Patrimoniale',
        sheet_stats='5b. Stat. Eq. Patrimoniale',
        chart_title='Andamento Mediano Indici',
        y_axis_name='Valori',
        rename_dict={
            'struttura 1° livello': 'Indice Struttura 1° Liv.',
            'struttura 2° livello': 'Indice Struttura 2° Liv.',
            'gearing': 'Gearing'
        },
        azienda_target=azienda_target
    )

    writer.close()
    output_buffer.seek(0)
    
    return output_buffer


def elabora_capitolo_5(df_filtered, azienda_target):
    import io
    import pandas as pd
    import numpy as np
    import xlsxwriter
    from xlsxwriter.utility import xl_rowcol_to_cell
    import re

    def costruisci_sezione_analisi(writer, workbook, df_raw, formati, keyword_ricerca, sheet_data, sheet_stats, chart_title, y_axis_name, rename_dict, azienda_target):
        base_cols = [c for c in df_raw.columns if 'ragione' in str(c).lower() or 'bvd' in str(c).lower() or 'nuts2' in str(c).lower() or 'nuts3' in str(c).lower()]
        
        if isinstance(keyword_ricerca, list):
            metric_cols = [c for c in df_raw.columns if any(kw.lower() in str(c).lower() for kw in keyword_ricerca)]
        else:
            metric_cols = [c for c in df_raw.columns if keyword_ricerca.lower() in str(c).lower()]

        if not metric_cols:
            return

        df = df_raw[base_cols + metric_cols].copy()
        df = df.replace(['n.d.', 'n.a.', 'n.s.', ''], np.nan)
        num_rows = len(df)

        col_ragione_idx = [idx for idx, c in enumerate(df.columns) if 'ragione' in str(c).lower()][0]
        col_ragione = df.columns[col_ragione_idx]

        df.to_excel(writer, sheet_name=sheet_data, index=False, startrow=0)
        worksheet_data = writer.sheets[sheet_data]
        worksheet_stats = workbook.add_worksheet(sheet_stats)

        # 3. Compilazione Foglio 1: Dati, Autofit ed EVIDENZIAZIONE AZIENDA TARGET
        fmt_target_text = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'left', 'valign': 'vcenter', 'bold': True})
        fmt_target_num = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'bold': True})

        for i, col in enumerate(df.columns):
            worksheet_data.write(0, i, col, formati['header'])
            col_data = df[col].dropna()
            max_len = len(str(col)) + 2 if col_data.empty else max(col_data.astype(str).map(len).max(), len(str(col))) + 2
            worksheet_data.set_column(i, i, min(max_len, 45))

            for row in range(1, num_rows + 1):
                val = df.iat[row-1, i]
                is_target = azienda_target.lower().strip() in str(df.iat[row-1, col_ragione_idx]).lower().strip()

                f_text = fmt_target_text if is_target else formati['data_text']
                f_num = fmt_target_num if is_target else formati['data_num']

                if pd.isna(val):
                    worksheet_data.write(row, i, "", f_text)
                elif isinstance(val, (int, float)):
                    worksheet_data.write_number(row, i, val, f_num)
                else:
                    worksheet_data.write_string(row, i, str(val), f_text)

        numeric_cols_idx = [df.columns.get_loc(c) for c in metric_cols]
        metrics_dict = {}

        for col_idx in numeric_cols_idx:
            col_name = df.columns[col_idx]
            anno_match = re.search(r'\d{4}', col_name)
            if anno_match:
                anno = int(anno_match.group(0))
                metric_base = re.sub(r'\d{4}', '', col_name).replace('(*) %', '').replace('(*)', '').replace('migl EUR', '').replace(' ', ' ').strip()
                
                for k, v in rename_dict.items():
                    if k.lower() in metric_base.lower():
                        metric_base = v

                if metric_base not in metrics_dict: metrics_dict[metric_base] = {}
                metrics_dict[metric_base][anno] = col_idx

        all_years = sorted(list(set(anno for m in metrics_dict for anno in metrics_dict[m].keys())))
        metrics_list = list(metrics_dict.keys())

        col_left = 0
        col_mid = len(all_years) + 2
        col_right = col_mid + (len(all_years)*2 - 1) + 2

        worksheet_stats.set_column(col_left, col_left, 22)
        worksheet_stats.set_column(col_left + 1, col_left + len(all_years), 12)
        worksheet_stats.set_column(col_mid, col_mid, 22)
        worksheet_stats.set_column(col_mid + 1, col_mid + (len(all_years)*2), 10)
        worksheet_stats.set_column(col_right, col_right, 22)
        worksheet_stats.set_column(col_right + 1, col_right + len(all_years), 12)

        median_cells = {}
        current_row = 1
        stats_formulas = [('Media', 'AVERAGE'), ('Mediana', 'MEDIAN'), ('Asimmetria', 'SKEW'), ('Curtosi', 'KURT'), ('Deviazione standard', 'STDEV')]

        for metric in metrics_list:
            median_cells[metric] = {}

            worksheet_stats.write(current_row, col_left, "", formati['header'])
            for i, year in enumerate(all_years):
                worksheet_stats.write(current_row, col_left + 1 + i, year, formati['header'])

            for s_idx, (stat_name, func) in enumerate(stats_formulas):
                r = current_row + 1 + s_idx
                worksheet_stats.write(r, col_left, stat_name, formati['label'])
                for i, year in enumerate(all_years):
                    if year in metrics_dict[metric]:
                        data_col_idx = metrics_dict[metric][year]
                        col_letter = xlsxwriter.utility.xl_col_to_name(data_col_idx)
                        data_range = f"'{sheet_data}'!{col_letter}2:{col_letter}{num_rows+1}"
                        formula = f"={func}({data_range})"
                        worksheet_stats.write_formula(r, col_left + 1 + i, formula, formati['data_num'])
                        if stat_name == 'Mediana':
                            median_cells[metric][year] = xl_rowcol_to_cell(r, col_left + 1 + i)
                    else:
                        worksheet_stats.write(r, col_left + 1 + i, "n.d.", formati['data_text'])

            c_offset = 1
            for i, year in enumerate(all_years):
                if i == 0:
                    worksheet_stats.write(current_row, col_mid + c_offset, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    c_offset += 1
                else:
                    worksheet_stats.merge_range(current_row, col_mid + c_offset, current_row, col_mid + c_offset + 1, year, formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset, "V.m.", formati['header'])
                    worksheet_stats.write(current_row + 1, col_mid + c_offset + 1, "Δ %", formati['header'])
                    c_offset += 2

            r_data = current_row + 2
            worksheet_stats.write(r_data, col_mid, metric, formati['label'])
            c_offset = 1
            prev_target_cell = None
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    source_cell = median_cells[metric][year]
                    curr_target_cell = xl_rowcol_to_cell(r_data, col_mid + c_offset)
                    worksheet_stats.write_formula(r_data, col_mid + c_offset, f"={source_cell}", formati['data_num'])

                    if i > 0:
                        if prev_target_cell:
                            delta_formula = f"=IFERROR(({curr_target_cell}-{prev_target_cell})/ABS({prev_target_cell}), 0)"
                            worksheet_stats.write_formula(r_data, col_mid + c_offset + 1, delta_formula, formati['pct'])
                        else:
                            worksheet_stats.write(r_data, col_mid + c_offset + 1, "n.d.", formati['data_text'])
                        prev_target_cell = curr_target_cell
                        c_offset += 2
                    else:
                        prev_target_cell = curr_target_cell
                        c_offset += 1
                else:
                    prev_target_cell = curr_target_cell
                    c_offset += 1

            current_row += 8 

        # 6. Costruzione Blocco Destro: CONFRONTO DIRETTO RIGHE SETTORE VS AZIENDA TARGET
        r_right = 1
        worksheet_stats.write(r_right, col_right, "Confronto Benchmark", formati['header'])
        for i, year in enumerate(all_years):
            worksheet_stats.write(r_right, col_right + 1 + i, year, formati['header'])

        df_azienda = df[df[col_ragione].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
        fmt_az_label = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'font_color': '#002060', 'border': 1, 'align': 'left'})
        fmt_az_num = workbook.add_format({'bg_color': '#F2F4F8', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'bold': True})

        row_cursor = r_right + 1
        for metric in metrics_list:
            # Riga 1: Settore
            worksheet_stats.write(row_cursor, col_right, f"{metric} (Settore)", formati['label'])
            for i, year in enumerate(all_years):
                if year in median_cells[metric]:
                    worksheet_stats.write_formula(row_cursor, col_right + 1 + i, f"={median_cells[metric][year]}", formati['data_num'])
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

            # Riga 2: Azienda Target
            worksheet_stats.write(row_cursor, col_right, f"{metric} ({azienda_target})", fmt_az_label)
            for i, year in enumerate(all_years):
                if year in metrics_dict[metric] and not df_azienda.empty:
                    val_az = df_azienda.iloc[0, metrics_dict[metric][year]]
                    if pd.isna(val_az) or val_az == "":
                        worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
                    else:
                        worksheet_stats.write_number(row_cursor, col_right + 1 + i, float(val_az), fmt_az_num)
                else:
                    worksheet_stats.write(row_cursor, col_right + 1 + i, "n.d.", formati['data_text'])
            row_cursor += 1

        # ==========================================
        # 7. INSERIMENTO GRAFICI DINAMICI UNIFORMI (CAP 3, 4, 5)
        # ==========================================
        row_start_metrics = r_right + 1
        start_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + 1)
        end_col_letter = xlsxwriter.utility.xl_col_to_name(col_right + len(all_years))
        last_year_col = end_col_letter 
        
        safe_chart_row = max(current_row, row_cursor)

        # 🟢 IMPOSTAZIONI ANTI-ACCAVALLAMENTO UNIVERSALI
        # Riduciamo i font per impedire che i numeri giganteschi sfondino i margini del grafico
        font_assi = {'size': 9}
        font_etichette = {'size': 8}
        
        # --- GRAFICO 1 (LINEE CUMULATIVE): Trend Settore ---
        chart_trend_settore = workbook.add_chart({'type': 'line'})
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            chart_trend_settore.add_series({
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}",
                'marker': {'type': 'circle'}
            })
        chart_trend_settore.set_title({'name': f'Andamento Storico di Settore ({chart_title.replace("Andamento Mediano ", "")})'})
        chart_trend_settore.set_legend({'position': 'bottom', 'font': font_assi})
        chart_trend_settore.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
        chart_trend_settore.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
        chart_trend_settore.set_size({'width': 650, 'height': 350})
        chart_trend_settore.set_style(11)
        worksheet_stats.insert_chart(safe_chart_row, col_left, chart_trend_settore)

        # --- GRAFICI MULTIPLI (COPPIE) ---
        chart_offset_y = safe_chart_row
        
        for m_idx, metric in enumerate(metrics_list):
            r_settore = row_start_metrics + (m_idx * 2)
            r_azienda = r_settore + 1
            
            # 🟢 FIX ISTOGRAMMA: Se la metrica riguarda grandi volumi, abbandona la linea e crea un istogramma standard
            is_volumi = any(k in metric.lower() for k in ['ricavi', 'attivo', 'produzione'])
            tipo_storico = 'column' if is_volumi else 'line'

            # A) Grafico Storico (Linee o Colonne in base alla metrica)
            chart_storico = workbook.add_chart({'type': tipo_storico})
            
            # Setup posizioni etichette: se istogramma vanno "sopra il tetto" della barra, se linea le separiamo (Sopra/Sotto)
            pos_label_settore = 'outside_end' if is_volumi else 'below'
            pos_label_azienda = 'outside_end' if is_volumi else 'above'

            # 🟢 Disattiviamo i numeri fluttuanti: i valori precisi saranno nella Tabella in basso
            serie_settore_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_settore + 1}:${end_col_letter}${r_settore + 1}"
            }
            serie_azienda_storico = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${start_col_letter}$2:${end_col_letter}$2",
                'values': f"='{sheet_stats}'!${start_col_letter}${r_azienda + 1}:${end_col_letter}${r_azienda + 1}"
            }

            # ... (codice marker rimane intatto) ...
            if not is_volumi:
                serie_settore_storico['marker'] = {'type': 'circle'}
                serie_azienda_storico['marker'] = {'type': 'square'}

            chart_storico.add_series(serie_azienda_storico)
            chart_storico.add_series(serie_settore_storico)
            chart_storico.set_title({'name': f'Benchmark Storico - {metric}'})

            # 🟢 Nessuna legenda necessaria: la tabella dati include già i nomi delle serie!
            chart_storico.set_legend({'none': True}) 

            # 🟢 ATTIVAZIONE MATRICE DATI (DATA TABLE)
            chart_storico.set_table({'show_keys': True})

            chart_storico.set_x_axis({'name_font': font_assi, 'num_font': font_assi})
            # 🟢 Togli la griglia orizzontale SOLO se il grafico storico è un istogramma (es. Ricavi)
            if is_volumi:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi, 'major_gridlines': {'visible': False}})
            else:
                chart_storico.set_y_axis({'name_font': font_assi, 'num_font': font_assi})
            chart_storico.set_size({'width': 550, 'height': 350}) 
            chart_storico.set_style(11)
            
            # B) Grafico a Colonne / Istogramma (Azienda vs Settore - Solo anno 2024)
            chart_col_singolo = workbook.add_chart({'type': 'column'})
            
            serie_settore_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_settore + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_settore + 1}:${last_year_col}${r_settore + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            serie_azienda_col = {
                'name': f"='{sheet_stats}'!${xlsxwriter.utility.xl_col_to_name(col_right)}${r_azienda + 1}",
                'categories': f"='{sheet_stats}'!${last_year_col}$2:${last_year_col}$2",
                'values': f"='{sheet_stats}'!${last_year_col}${r_azienda + 1}:${last_year_col}${r_azienda + 1}",
                'data_labels': {'value': True, 'num_format': '#,##0.00', 'font': font_etichette, 'position': 'outside_end'} 
            }
            chart_col_singolo.add_series(serie_azienda_col)
            chart_col_singolo.add_series(serie_settore_col)
            chart_col_singolo.set_title({'name': f'Posizionamento {metric} (Anno 2024)'})
            chart_col_singolo.set_legend({'position': 'bottom', 'font': font_assi})
            chart_col_singolo.set_x_axis({'name_font': font_assi, 'num_font': font_assi, 'label_position': 'low'})
            # 🟢 Rimuove la griglia orizzontale di sfondo dall'istogramma
            chart_col_singolo.set_y_axis({
                'name_font': font_assi, 
                'num_font': font_assi, 
                'major_gridlines': {'visible': False}
            })
            chart_col_singolo.set_size({'width': 350, 'height': 350})
            chart_col_singolo.set_style(11)

            # Inseriamo i grafici distanziandoli a dovere sulla destra
            worksheet_stats.insert_chart(chart_offset_y, col_right + 1, chart_storico)
            worksheet_stats.insert_chart(chart_offset_y, col_right + 9, chart_col_singolo)
            
            # Incremento dello step verticale da 16 a 18 righe per distanziare nettamente una metrica dall'altra
            chart_offset_y += 20 

        worksheet_stats.ignore_errors({'formula_differs': 'A1:Z500', 'number_stored_as_text': 'A1:Z500'})

    df_raw = df_filtered.copy()
    output_buffer = io.BytesIO()
    
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book

    formati = {
        'header': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter'}),
        'label': workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'data_num': workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter'}),
        'data_text': workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'}),
        'pct': workbook.add_format({'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})
    }

    costruisci_sezione_analisi(
        writer, workbook, df_raw, formati,
        keyword_ricerca=['current ratio', 'quick ratio', 'rotazione'], 
        sheet_data='6a. Dati Eq. Finanziario',
        sheet_stats='6b. Stat. Eq. Finanziario',
        chart_title='Andamento Mediano Indici',
        y_axis_name='Valori',
        rename_dict={
            'current ratio': 'Current Ratio',
            'quick ratio': 'Quick Ratio',
            'rotazione del capitale investito': 'Indice Rotazione Cap.Inv.'
        },
        azienda_target=azienda_target
    )

    writer.close()
    output_buffer.seek(0)
    
    return output_buffer


def elabora_capitolo_6(df_filtered, azienda_target):
    import io
    import pandas as pd
    import numpy as np
    import xlsxwriter

    # 2. Caricamento dati (in memory)
    df_raw = df_filtered.copy()
    
    # 3. Definizione colonne
    def trova_col(keywords, exclude=None):
        if exclude is None: exclude = []
        for c in df_raw.columns:
            c_lower = str(c).lower()
            if all(k.lower() in c_lower for k in keywords) and not any(e.lower() in c_lower for e in exclude):
                return c
        return None

    col_regione = trova_col(['nuts', '2'])
    if not col_regione: col_regione = trova_col(['nuts'])
    
    col_macro = trova_col(['nuts', '1'])
    if not col_macro:
        if col_regione:
            df_raw['Macro_Fallback'] = df_raw[col_regione].astype(str).str[:3] + " - Macroregione"
            col_macro = 'Macro_Fallback'
        else:
             df_raw['Macro_Fallback'] = "N.D. - Macroregione"
             col_macro = 'Macro_Fallback'

    cols_dict = {
        'Ragione Sociale': trova_col(['ragione']),
        'Macro-Regione': col_macro, 
        'Regione': col_regione,
        'M. Profitto 2024': trova_col(['marg', 'profitto', '2024']),
        'M. EBITDA 2024': trova_col(['marg', 'ebitda', '2024']),
        'M. EBIT 2024': trova_col(['marg', 'ebit', '2024'], exclude=['ebitda']),
        'Rotazione C.Inv. 2024': trova_col(['rotazione', '2024']),
        'Quick Ratio 2024': trova_col(['quick', '2024']),
        'Current Ratio 2024': trova_col(['current', '2024']),
        'Indice 1° Liv. 2024': trova_col(['struttura 1', '2024']),
        'Indice 2° Liv. 2024': trova_col(['struttura 2', '2024']),
        'Gearing 2024': trova_col(['gearing', '2024'])
    }

    if cols_dict['Regione'] is None:
        df_raw['Regione_Fallback'] = "N.D."
        cols_dict['Regione'] = 'Regione_Fallback'

    df = df_raw[list(cols_dict.values())].copy()
    df.columns = list(cols_dict.keys())
    
    metriche = list(cols_dict.keys())[3:] 
    for m in metriche:
        df[m] = pd.to_numeric(df[m], errors='coerce')

    # 4. Creazione Colonne Vuote per Spaziature e Formule
    df.insert(1, 'Società ID', range(1, len(df) + 1))
    
    df[' '] = ''    
    df['  '] = ''   
    df['   '] = ''  
    df['    '] = '' 

    df['Benchmark Economico'] = ''
    df['Benchmark Finanziario'] = ''
    df['Benchmark Patrimoniale'] = ''
    df['Benchmark Totale'] = ''
    df['Rating Combinato'] = ''

    # 5. Ordinamento colonne
    col_finali = [
        'Ragione Sociale', 'Società ID',
        ' ', 
        'M. Profitto 2024', 'M. EBITDA 2024', 'M. EBIT 2024',
        '  ', 
        'Rotazione C.Inv. 2024', 'Quick Ratio 2024', 'Current Ratio 2024',
        '   ', 
        'Indice 1° Liv. 2024', 'Indice 2° Liv. 2024', 'Gearing 2024',
        '    ', 
        'Benchmark Economico', 'Benchmark Finanziario', 'Benchmark Patrimoniale', 'Benchmark Totale',
        'Regione', 'Rating Combinato'
    ]
    df_out = df[col_finali]

    start_col_tbl = len(df_out.columns) + 1 
    col_T1 = xlsxwriter.utility.xl_col_to_name(start_col_tbl + 2) 
    col_T2 = xlsxwriter.utility.xl_col_to_name(start_col_tbl + 3) 

    output_buffer = io.BytesIO()
    
    # 6. Scrittura su Excel (Foglio 7)
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book
    worksheet = workbook.add_worksheet('7. Rating e Benchmark 2024')
    
    fmt_header = workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_header_metric = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_data = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter'})
    fmt_num = workbook.add_format({'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter'})
    fmt_center = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
    fmt_id = workbook.add_format({'border': 1, 'num_format': '0', 'align': 'center', 'valign': 'vcenter'}) 
    fmt_space = workbook.add_format({'bg_color': '#EAEAEA'}) 

    for col_num, col_name in enumerate(df_out.columns):
        if col_name.strip() == '':
            worksheet.write(0, col_num, "", fmt_space)
        else:
            formato = fmt_header_metric if '2024' in col_name else fmt_header
            worksheet.write(0, col_num, col_name, formato)
        
    # Creazione degli stili di evidenziazione per l'azienda target nella lista
    fmt_target_data = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'left', 'valign': 'vcenter', 'bold': True})
    fmt_target_num = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '#,##0.00', 'align': 'right', 'valign': 'vcenter', 'bold': True})
    fmt_target_center = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True})

    idx_target_salvata = None # Ci servirà per rintracciare la riga esatta e clonarla sopra

    for row_num, row_data in enumerate(df_out.values):
        row_ex = row_num + 2
        # Verifica se la riga corrente appartiene all'azienda target
        is_target_row = azienda_target.lower().strip() in str(row_data[0]).lower().strip()

        if is_target_row:
            idx_target_salvata = row_ex # Salva l'indice di riga excel attuale

        # Sostituiamo i formati standard con quelli target se is_target_row è True
        f_dat = fmt_target_data if is_target_row else fmt_data
        f_nm = fmt_target_num if is_target_row else fmt_num
        f_cntr = fmt_target_center if is_target_row else fmt_center

        for col_num, val in enumerate(row_data):
            col_name = df_out.columns[col_num]
            if col_name.strip() == '':
                worksheet.write(row_ex - 1, col_num, "", fmt_space)
            elif col_name == 'Benchmark Economico':
                # Riferimenti corretti: Profitto riga 8, EBITDA riga 9, EBIT riga 10
                cond_D = f"IF(D{row_ex}>=${col_T2}$8,3,IF(D{row_ex}>=${col_T1}$8,2,1))"
                cond_E = f"IF(E{row_ex}>=${col_T2}$9,3,IF(E{row_ex}>=${col_T1}$9,2,1))"
                cond_F = f"IF(F{row_ex}>=${col_T2}$10,3,IF(F{row_ex}>=${col_T1}$10,2,1))"
                formula = f'=IF(({cond_D}+{cond_E}+{cond_F})>=8,"A",IF(({cond_D}+{cond_E}+{cond_F})>=5,"B","C"))'
                worksheet.write_formula(row_ex - 1, col_num, formula, f_cntr)
            elif col_name == 'Benchmark Finanziario':
                # Riferimenti corretti: Rotazione riga 11, Quick riga 12, Current riga 13
                cond_H = f"IF(H{row_ex}>=${col_T2}$11,3,IF(H{row_ex}>=${col_T1}$11,2,1))"
                cond_I = f"IF(I{row_ex}>=${col_T2}$12,3,IF(I{row_ex}>=${col_T1}$12,2,1))"
                cond_J = f"IF(J{row_ex}>=${col_T2}$13,3,IF(J{row_ex}>=${col_T1}$13,2,1))"
                formula = f'=IF(({cond_H}+{cond_I}+{cond_J})>=8,"A",IF(({cond_H}+{cond_I}+{cond_J})>=5,"B","C"))'
                worksheet.write_formula(row_ex - 1, col_num, formula, f_cntr)
            elif col_name == 'Benchmark Patrimoniale':
                # Riferimenti corretti: Struttura 1° riga 14, Struttura 2° riga 15, Gearing riga 16
                cond_L = f"IF(L{row_ex}>=${col_T2}$14,3,IF(L{row_ex}>=${col_T1}$14,2,1))"
                cond_M = f"IF(M{row_ex}>=${col_T2}$15,3,IF(M{row_ex}>=${col_T1}$15,2,1))"
                cond_N = f"IF(N{row_ex}<=${col_T1}$16,3,IF(N{row_ex}<=${col_T2}$16,2,1))"
                formula = f'=IF(({cond_L}+{cond_M}+{cond_N})>=8,"A",IF(({cond_L}+{cond_M}+{cond_N})>=5,"B","C"))'
                worksheet.write_formula(row_ex - 1, col_num, formula, f_cntr)
            elif col_name == 'Benchmark Totale':
                cond_P = f'IF(P{row_ex}="A",3,IF(P{row_ex}="B",2,1))'
                cond_Q = f'IF(Q{row_ex}="A",3,IF(Q{row_ex}="B",2,1))'
                cond_R = f'IF(R{row_ex}="A",3,IF(R{row_ex}="B",2,1))'
                formula = f'=IF(({cond_P}+{cond_Q}+{cond_R})>=8,"A",IF(({cond_P}+{cond_Q}+{cond_R})>=5,"B","C"))'
                worksheet.write_formula(row_ex - 1, col_num, formula, f_cntr)
            elif col_name == 'Rating Combinato':
                # 🟢 FIX: Ordine istituzionale (Economico + Patrimoniale + Finanziario)
                formula = f'=P{row_ex}&R{row_ex}&Q{row_ex}'
                worksheet.write_formula(row_ex - 1, col_num, formula, f_cntr)
            elif pd.isna(val):
                worksheet.write(row_ex - 1, col_num, "n.d.", f_dat)
            elif col_name == 'Società ID':
                worksheet.write_number(row_ex - 1, col_num, val, fmt_id)
            elif isinstance(val, (int, float)):
                worksheet.write_number(row_ex - 1, col_num, val, f_nm)
            else:
                worksheet.write(row_ex - 1, col_num, str(val), f_dat)

    for col_num, col_name in enumerate(df_out.columns):
        if col_name.strip() == '':
            worksheet.set_column(col_num, col_num, 2) 
            
    worksheet.set_column('A:A', 35)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('D:F', 14) 
    worksheet.set_column('H:J', 14) 
    worksheet.set_column('L:N', 14) 
    worksheet.set_column('P:S', 20) 
    worksheet.set_column('T:T', 35) 
    worksheet.set_column('U:U', 18) 
    worksheet.set_row(0, 45) 

    # Se l'azienda è stata rintracciata, clona le sue metriche fondamentali in un pannello superiore isolato
    if idx_target_salvata is not None:
        worksheet.write(1, start_col_tbl, "RATING ISOLATO AZIENDA TARGET", fmt_header)
        worksheet.write(2, start_col_tbl, "Ragione Sociale", fmt_header_metric)
        worksheet.write(2, start_col_tbl + 1, "Rating Finale", fmt_header_metric)
        worksheet.write(2, start_col_tbl + 2, "Combinazione", fmt_header_metric)

        # Formule collegate alla riga reale per garantire coerenza matematica
        worksheet.write_formula(3, start_col_tbl, f"=A{idx_target_salvata}", fmt_data)
        worksheet.write_formula(3, start_col_tbl + 1, f"=S{idx_target_salvata}", fmt_center)
        worksheet.write_formula(3, start_col_tbl + 2, f"=U{idx_target_salvata}", fmt_center)

    # Scrittura standard delle soglie spostata più in basso (riga 6 anziché riga 1) per fare spazio al box sopra
    worksheet.write(5, start_col_tbl, "SOGLIE CALCOLATE 2024", fmt_header)
    worksheet.write(6, start_col_tbl, "Metrica", fmt_header)
    intestazioni_tbl = ["MIN", "Soglia 2° Terzile", "Soglia 1° Terzile", "MAX"]
    for i, h in enumerate(intestazioni_tbl):
        worksheet.write(6, start_col_tbl + 1 + i, h, fmt_header)

    num_rows = len(df_out)
    for i, m in enumerate(metriche):
        riga = 7 + i  # Scalata di 4 righe in basso
        worksheet.write(riga, start_col_tbl, m, fmt_header_metric)
        col_idx = df_out.columns.get_loc(m)
        col_letter = xlsxwriter.utility.xl_col_to_name(col_idx)
        data_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
        worksheet.write_formula(riga, start_col_tbl + 1, f"=MIN({data_range})", fmt_num)
        worksheet.write_formula(riga, start_col_tbl + 2, f"=PERCENTILE({data_range}, 1/3)", fmt_num)
        worksheet.write_formula(riga, start_col_tbl + 3, f"=PERCENTILE({data_range}, 2/3)", fmt_num)
        worksheet.write_formula(riga, start_col_tbl + 4, f"=MAX({data_range})", fmt_num)
        
    worksheet.set_column(start_col_tbl, start_col_tbl, 20)
    worksheet.set_column(start_col_tbl + 1, start_col_tbl + 4, 18)

    # ---------------------------------------------------------
    # AGGIUNTA FOGLI PIVOT E RIPARTIZIONE MACRO/NUTS2
    # ---------------------------------------------------------
    fmt_pct = workbook.add_format({'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})
    fmt_num_int = workbook.add_format({'border': 1, 'num_format': '#,##0', 'align': 'right', 'valign': 'vcenter'})
    
    fmt_subtotal = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'border': 1, 'align': 'left', 'valign': 'vcenter'})
    fmt_subtotal_num = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'border': 1, 'num_format': '#,##0', 'align': 'right', 'valign': 'vcenter'})
    fmt_subtotal_pct = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})
    fmt_grand_tot = workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'align': 'left', 'valign': 'vcenter'})
    fmt_grand_tot_num = workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'num_format': '#,##0', 'align': 'right', 'valign': 'vcenter'})
    fmt_grand_tot_pct = workbook.add_format({'bold': True, 'bg_color': '#002060', 'font_color': 'white', 'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter'})

    bench_cols = {
        'Benchmark Economico': 'P',
        'Benchmark Finanziario': 'Q',
        'Benchmark Patrimoniale': 'R',
        'Benchmark Totale': 'S'
    }
    main_sheet = "'7. Rating e Benchmark 2024'"
    range_reg = f"{main_sheet}!$T$2:$T${num_rows+1}"

    # --- Foglio 8: Pivot Analisi Benchmark ---
    worksheet_pivot = workbook.add_worksheet('8. Pivot Analisi Benchmark')
    worksheet_pivot.set_column('A:A', 35)
    worksheet_pivot.set_column('B:C', 15)

    row_p = 1
    for b_name, b_col in bench_cols.items():
        worksheet_pivot.write(row_p, 0, f"Conteggio di {b_name}", fmt_header)
        worksheet_pivot.write(row_p, 1, "V.A.", fmt_header)
        worksheet_pivot.write(row_p, 2, "%", fmt_header)
        row_p += 1

        range_bench = f"{main_sheet}!${b_col}$2:${b_col}${num_rows+1}"
        row_start = row_p

        for lettera in ['A', 'B', 'C']:
            worksheet_pivot.write(row_p, 0, lettera, fmt_center)
            worksheet_pivot.write_formula(row_p, 1, f'=COUNTIF({range_bench}, "{lettera}")', fmt_num_int)
            row_p += 1

        row_tot = row_p
        worksheet_pivot.write(row_tot, 0, "Totale complessivo", fmt_header)
        worksheet_pivot.write_formula(row_tot, 1, f'=SUM(B{row_start+1}:B{row_tot})', fmt_num_int)
        worksheet_pivot.write(row_tot, 2, 1.0, fmt_pct)

        for i in range(3):
            worksheet_pivot.write_formula(row_start + i, 2, f'=IF($B${row_tot+1}>0, B{row_start+i+1}/$B${row_tot+1}, 0)', fmt_pct)

        row_p += 3

    # --- Foglio 9: Rip.Terr. Benchmark ---
    worksheet_terr = workbook.add_worksheet('9. Rip.Terr. Benchmark')
    worksheet_terr.set_column('A:A', 40)
    worksheet_terr.set_column('B:I', 15)

    # 🟢 ESTRATTO TARGET CAP 6: Trova Regione e Macroregione
    df_az_geo_check = df[df['Ragione Sociale'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    target_reg_cap6 = df_az_geo_check.iloc[0]['Regione'] if not df_az_geo_check.empty else None
    target_mac_cap6 = df_az_geo_check.iloc[0]['Macro-Regione'] if not df_az_geo_check.empty else None

    # Stili Evidenziazione Gialla per xlsxwriter
    fmt_hl_data = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'align': 'left', 'valign': 'vcenter', 'bold': True})
    fmt_hl_num  = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '#,##0', 'align': 'right', 'valign': 'vcenter', 'bold': True})
    fmt_hl_pct  = workbook.add_format({'bg_color': '#FFF2CC', 'border': 1, 'num_format': '0.00%', 'align': 'right', 'valign': 'vcenter', 'bold': True})

    mapping_df = df[['Macro-Regione', 'Regione']].dropna()
    mapping_df = mapping_df[(mapping_df['Regione'].astype(str).str.strip() != '') & (mapping_df['Regione'].astype(str).str.lower() != 'nan')]
    
    mapping = {}
    for macro, group in mapping_df.groupby('Macro-Regione'):
        macro_str = str(macro).strip()
        if macro_str and macro_str.lower() != 'nan':
            mapping[macro_str] = group['Regione'].unique().tolist()

    sorted_macro = sorted(mapping.keys())

    row_t = 1
    for b_name, b_col in bench_cols.items():
        range_bench = f"{main_sheet}!${b_col}$2:${b_col}${num_rows+1}"

        worksheet_terr.merge_range(row_t, 0, row_t, 8, f"{b_name} - Ripartizione Territoriale", fmt_header_metric)
        row_t += 1

        headers = ["Regioni", "Imprese V.A.", "Imprese %", "A (V.A.)", "A (%)", "B (V.A.)", "B (%)", "C (V.A.)", "C (%)"]
        for i, h in enumerate(headers):
            worksheet_terr.write(row_t, i, h, fmt_header)
        row_t += 1

        offset_righe = sum(len(mapping[m]) + 2 for m in sorted_macro) - 1 if sorted_macro else 0
        riga_italia = row_t + offset_righe
        total_cell = f"$B${riga_italia+1}"
        
        grand_total_row_refs = []

        for macro in sorted_macro:
            regioni_list = mapping[macro]
            start_macro_row = row_t
            
            # --- CICLO DELLE SINGOLE REGIONI ---
            for reg in sorted(regioni_list):
                # Controlla se questa regione è quella dell'azienda target
                is_reg_target = (target_reg_cap6 is not None and str(reg).lower().strip() == str(target_reg_cap6).lower().strip())
                
                # Seleziona il formato (Giallo se target, altrimenti standard)
                f_dat = fmt_hl_data if is_reg_target else fmt_data
                f_num = fmt_hl_num if is_reg_target else fmt_num_int
                f_pct = fmt_hl_pct if is_reg_target else fmt_pct

                worksheet_terr.write(row_t, 0, str(reg), f_dat)
                worksheet_terr.write_formula(row_t, 1, f'=COUNTIF({range_reg}, "{reg}")', f_num)
                worksheet_terr.write_formula(row_t, 2, f'=IF({total_cell}>0, B{row_t+1}/{total_cell}, 0)', f_pct)
                worksheet_terr.write_formula(row_t, 3, f'=COUNTIFS({range_reg}, "{reg}", {range_bench}, "A")', f_num)
                worksheet_terr.write_formula(row_t, 4, f'=IF({total_cell}>0, D{row_t+1}/{total_cell}, 0)', f_pct)
                worksheet_terr.write_formula(row_t, 5, f'=COUNTIFS({range_reg}, "{reg}", {range_bench}, "B")', f_num)
                worksheet_terr.write_formula(row_t, 6, f'=IF({total_cell}>0, F{row_t+1}/{total_cell}, 0)', f_pct)
                worksheet_terr.write_formula(row_t, 7, f'=COUNTIFS({range_reg}, "{reg}", {range_bench}, "C")', f_num)
                worksheet_terr.write_formula(row_t, 8, f'=IF({total_cell}>0, H{row_t+1}/{total_cell}, 0)', f_pct)
                row_t += 1
                
            # --- RIGA TOTALE MACROREGIONE ---
            is_mac_target = (target_mac_cap6 is not None and str(macro).lower().strip() == str(target_mac_cap6).lower().strip())
            
            f_sub_dat = fmt_hl_data if is_mac_target else fmt_subtotal
            f_sub_num = fmt_hl_num if is_mac_target else fmt_subtotal_num
            f_sub_pct = fmt_hl_pct if is_mac_target else fmt_subtotal_pct

            worksheet_terr.write(row_t, 0, str(macro), f_sub_dat)
            worksheet_terr.write_formula(row_t, 1, f'=SUM(B{start_macro_row+1}:B{row_t})', f_sub_num)
            worksheet_terr.write_formula(row_t, 2, f'=IF({total_cell}>0, B{row_t+1}/{total_cell}, 0)', f_sub_pct)
            worksheet_terr.write_formula(row_t, 3, f'=SUM(D{start_macro_row+1}:D{row_t})', f_sub_num)
            worksheet_terr.write_formula(row_t, 4, f'=IF({total_cell}>0, D{row_t+1}/{total_cell}, 0)', f_sub_pct)
            worksheet_terr.write_formula(row_t, 5, f'=SUM(F{start_macro_row+1}:F{row_t})', f_sub_num)
            worksheet_terr.write_formula(row_t, 6, f'=IF({total_cell}>0, F{row_t+1}/{total_cell}, 0)', f_sub_pct)
            worksheet_terr.write_formula(row_t, 7, f'=SUM(H{start_macro_row+1}:H{row_t})', f_sub_num)
            worksheet_terr.write_formula(row_t, 8, f'=IF({total_cell}>0, H{row_t+1}/{total_cell}, 0)', f_sub_pct)
            
            grand_total_row_refs.append(row_t + 1)
            row_t += 2

        row_tot = row_t - 1 

        worksheet_terr.write(row_tot, 0, "Italia", fmt_grand_tot)
        if grand_total_row_refs:
            sum_b = "+".join([f"B{r}" for r in grand_total_row_refs])
            sum_d = "+".join([f"D{r}" for r in grand_total_row_refs])
            sum_f = "+".join([f"F{r}" for r in grand_total_row_refs])
            sum_h = "+".join([f"H{r}" for r in grand_total_row_refs])
            worksheet_terr.write_formula(row_tot, 1, f'={sum_b}', fmt_grand_tot_num)
            worksheet_terr.write_formula(row_tot, 3, f'={sum_d}', fmt_grand_tot_num)
            worksheet_terr.write_formula(row_tot, 5, f'={sum_f}', fmt_grand_tot_num)
            worksheet_terr.write_formula(row_tot, 7, f'={sum_h}', fmt_grand_tot_num)
        else:
             worksheet_terr.write(row_tot, 1, 0, fmt_grand_tot_num)
             worksheet_terr.write(row_tot, 3, 0, fmt_grand_tot_num)
             worksheet_terr.write(row_tot, 5, 0, fmt_grand_tot_num)
             worksheet_terr.write(row_tot, 7, 0, fmt_grand_tot_num)

        worksheet_terr.write_formula(row_tot, 2, f'=IF({total_cell}>0, B{row_tot+1}/{total_cell}, 0)', fmt_grand_tot_pct)
        worksheet_terr.write_formula(row_tot, 4, f'=IF({total_cell}>0, D{row_tot+1}/{total_cell}, 0)', fmt_grand_tot_pct)
        worksheet_terr.write_formula(row_tot, 6, f'=IF({total_cell}>0, F{row_tot+1}/{total_cell}, 0)', fmt_grand_tot_pct)
        worksheet_terr.write_formula(row_tot, 8, f'=IF({total_cell}>0, H{row_tot+1}/{total_cell}, 0)', fmt_grand_tot_pct)

        row_t += 4

    worksheet_terr.ignore_errors({'formula_differs': 'A1:Z1000'})

    writer.close()
    output_buffer.seek(0)
    
    return output_buffer



# INDICI DI COMPOSIZIONE

def elabora_capitolo_7_5(df_input, azienda_target):
    import io
    import pandas as pd
    import xlsxwriter

    df_base = df_input.copy()
    all_years = ['2021', '2022', '2023', '2024']

    componenti_nomi = {
        'Costo del venduto': 'Costo del venduto migl EUR',
        'Oneri diversi di gestione': 'Oneri diversi di gestione migl EUR',
        'Proventi/oneri finanziari': 'Proventi/oneri finanziari migl EUR',
        'Totale imposte': 'Totale imposte migl EUR',
        'Utile netto': 'Utile/Perdita al netto delle imposte migl EUR'
    }
    col_prod_prefisso = 'Totale valore della produzione migl EUR'

    for comp_label, col_prefix in componenti_nomi.items():
        for anno in all_years:
            col_name = f"{col_prefix} {anno}"
            if col_name in df_base.columns:
                df_base[col_name] = pd.to_numeric(df_base[col_name], errors='coerce').fillna(0)
    for anno in all_years:
        col_name = f"{col_prod_prefisso} {anno}"
        if col_name in df_base.columns:
            df_base[col_name] = pd.to_numeric(df_base[col_name], errors='coerce').fillna(1)

    df_base['is_target'] = df_base['Ragione socialeCaratteri latini'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)
    df_base = df_base.sort_values(by='is_target', ascending=False).reset_index(drop=True)
    df_base = df_base.drop(columns=['is_target'])

    output_buffer = io.BytesIO()
    writer = pd.ExcelWriter(output_buffer, engine='xlsxwriter')
    workbook = writer.book

    fmt_header = workbook.add_format({'bold': True, 'font_color': 'white', 'fg_color': '#4F81BD', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    fmt_text = workbook.add_format({'border': 1, 'align': 'left'})
    fmt_data_raw = workbook.add_format({'num_format': '#,##0', 'border': 1})
    fmt_data_pct = workbook.add_format({'num_format': '0.00%', 'border': 1, 'align': 'right'})
    fmt_data_tot = workbook.add_format({'num_format': '0.00%', 'border': 1, 'align': 'right', 'bold': True, 'fg_color': '#E2EFDA'})

    target_color = '#FFF2CC'
    fmt_bold_text = workbook.add_format({'bold': True, 'border': 1, 'fg_color': target_color})
    fmt_data_raw_tgt = workbook.add_format({'num_format': '#,##0', 'border': 1, 'bold': True, 'fg_color': target_color})
    fmt_data_pct_tgt = workbook.add_format({'num_format': '0.00%', 'border': 1, 'align': 'right', 'bold': True, 'fg_color': target_color})
    fmt_data_tot_tgt = workbook.add_format({'num_format': '0.00%', 'border': 1, 'align': 'right', 'bold': True, 'fg_color': '#FFE699'})

    # FOGLIO A: TABELLA DATI
    ws_dati = workbook.add_worksheet('7a. Dati Indici Composizione')
    ws_dati.hide_gridlines(2)
    ws_dati.set_column('A:A', 35)
    ws_dati.set_column('B:B', 15)
    ws_dati.set_column('C:ZZ', 14)

    ws_dati.write(0, 0, 'Ragione Sociale', fmt_header)
    ws_dati.write(0, 1, 'Numero BvD ID', fmt_header)

    col_idx = 2
    for anno in all_years: ws_dati.write(0, col_idx, f'Val. Prod. {anno}', fmt_header); col_idx += 1
    for comp in componenti_nomi.keys():
        for anno in all_years: ws_dati.write(0, col_idx, f'{comp} {anno}', fmt_header); col_idx += 1
    for comp in componenti_nomi.keys():
        for anno in all_years: ws_dati.write(0, col_idx, f'% {comp} {anno}', fmt_header); col_idx += 1
    for anno in all_years: ws_dati.write(0, col_idx, f'% TOTALE {anno}', fmt_header); col_idx += 1

    row_cursor_dati = 1
    for idx, row in df_base.iterrows():
        rag_soc = str(row.get('Ragione socialeCaratteri latini', ''))
        bvd_id = str(row.get('Numero BvD ID', ''))

        is_target = azienda_target.lower().strip() in rag_soc.lower()
        f_t = fmt_bold_text if is_target else fmt_text
        f_raw = fmt_data_raw_tgt if is_target else fmt_data_raw
        f_pct = fmt_data_pct_tgt if is_target else fmt_data_pct
        f_tot = fmt_data_tot_tgt if is_target else fmt_data_tot

        ws_dati.write(row_cursor_dati, 0, rag_soc, f_t)
        ws_dati.write(row_cursor_dati, 1, bvd_id, f_t)

        c_idx = 2
        
        # Salviamo le coordinate delle colonne per le formule
        prod_cols_map = {}
        for anno in all_years:
            v = row.get(f"{col_prod_prefisso} {anno}", 1)
            ws_dati.write(row_cursor_dati, c_idx, v, f_raw)
            prod_cols_map[anno] = xlsxwriter.utility.xl_col_to_name(c_idx)
            c_idx += 1

        abs_cols_map = {comp: {} for comp in componenti_nomi.keys()}
        for comp, col_prefix in componenti_nomi.items():
            for anno in all_years:
                v = row.get(f"{col_prefix} {anno}", 0)
                ws_dati.write(row_cursor_dati, c_idx, v, f_raw)
                abs_cols_map[comp][anno] = xlsxwriter.utility.xl_col_to_name(c_idx)
                c_idx += 1

        pct_cols_per_year = {anno: [] for anno in all_years}

        # 🟢 NUOVO: Formule Excel native per calcolare le percentuali (=Costo/ValoreProduzione)
        for comp, col_prefix in componenti_nomi.items():
            for anno in all_years:
                num_col = abs_cols_map[comp][anno]
                den_col = prod_cols_map[anno]
                r_num = row_cursor_dati + 1
                
                # Formula con controllo anti-divisione per zero (es: =IF(C2=0, 0, G2/C2))
                formula_pct = f"=IF({den_col}{r_num}=0, 0, {num_col}{r_num}/{den_col}{r_num})"
                ws_dati.write_formula(row_cursor_dati, c_idx, formula_pct, f_pct)
                
                pct_cols_per_year[anno].append((comp, xlsxwriter.utility.xl_col_to_name(c_idx)))
                c_idx += 1

        for anno in all_years:
            formula_parts = []
            for comp_name, col_letter in pct_cols_per_year[anno]:
                cell_ref = f"{col_letter}{row_cursor_dati+1}"
                if comp_name == 'Proventi/oneri finanziari':
                    formula_parts.append(f"-{cell_ref}")
                else:
                    formula_parts.append(f"+{cell_ref}")
            
            formula_somma = "=" + "".join(formula_parts).lstrip("+")
            ws_dati.write_formula(row_cursor_dati, c_idx, formula_somma, f_tot)
            c_idx += 1

        row_cursor_dati += 1

    # FOGLIO B: STATISTICHE E GRAFICI
    ws_stats = workbook.add_worksheet('7b. Stat. Indici Composizione')
    ws_stats.hide_gridlines(2)
    ws_stats.write(0, 0, 'Metrica / Componente %', fmt_header)
    ws_stats.set_column('A:A', 35)
    
    for c_idx, anno in enumerate(all_years, start=1):
        ws_stats.write(0, c_idx, anno, fmt_header)

    row_cursor = 2
    df_target_only = df_base[df_base['Ragione socialeCaratteri latini'].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    posizioni_grafici = {}

    for comp, col_prefix in componenti_nomi.items():
        ws_stats.write(row_cursor, 0, f"{comp} (Mediane Settore)", workbook.add_format({'bold': True, 'fg_color': '#DCE6F1', 'border': 1}))
        ws_stats.write(row_cursor + 1, 0, f"{comp} ({azienda_target})", workbook.add_format({'italic': True, 'border': 1}))

        posizioni_grafici[comp] = {'settore_row': row_cursor, 'azienda_row': row_cursor + 1}

        for c_idx, anno in enumerate(all_years, start=1):
            pct_series = (df_base[f"{col_prefix} {anno}"] / df_base[f"{col_prod_prefisso} {anno}"]).replace([float('inf'), -float('inf')], pd.NA).dropna()
            mediana_pct = pct_series.median() if not pct_series.empty else 0
            ws_stats.write(row_cursor, c_idx, mediana_pct, fmt_data_pct)

            tgt_pct = 0
            if not df_target_only.empty:
                v_prod_tgt = df_target_only.iloc[0].get(f"{col_prod_prefisso} {anno}", 1)
                if v_prod_tgt == 0: v_prod_tgt = 1
                tgt_pct = df_target_only.iloc[0].get(f"{col_prefix} {anno}", 0) / v_prod_tgt
            ws_stats.write(row_cursor + 1, c_idx, tgt_pct, fmt_data_pct)

        row_cursor += 3

    chart_offset_y = row_cursor + 4
    font_assi = {'size': 9}

    chart_settore = workbook.add_chart({'type': 'column', 'subtype': 'stacked'})
    chart_settore.set_title({'name': 'Composizione Storica - MEDIANE SETTORE'})
    chart_settore.set_table({'show_keys': True})
    chart_settore.set_legend({'none': True}) 
    chart_settore.set_y_axis({'max': 1, 'major_gridlines': {'visible': False}})
    chart_settore.set_size({'width': 550, 'height': 420}) 

    chart_azienda = workbook.add_chart({'type': 'column', 'subtype': 'stacked'})
    chart_azienda.set_title({'name': f'Composizione Storica - {azienda_target}'})
    chart_azienda.set_table({'show_keys': True})
    chart_azienda.set_legend({'none': True})
    chart_azienda.set_y_axis({'max': 1, 'major_gridlines': {'visible': False}})
    chart_azienda.set_size({'width': 550, 'height': 420})

    for idx, comp in enumerate(componenti_nomi.keys()):
        pos = posizioni_grafici[comp]
        chart_settore.add_series({
            'name': ['7b. Stat. Indici Composizione', pos['settore_row'], 0],
            'categories': ['7b. Stat. Indici Composizione', 0, 1, 0, len(all_years)],
            'values': ['7b. Stat. Indici Composizione', pos['settore_row'], 1, pos['settore_row'], len(all_years)],
        })
        chart_azienda.add_series({
            'name': ['7b. Stat. Indici Composizione', pos['azienda_row'], 0],
            'categories': ['7b. Stat. Indici Composizione', 0, 1, 0, len(all_years)],
            'values': ['7b. Stat. Indici Composizione', pos['azienda_row'], 1, pos['azienda_row'], len(all_years)],
        })

    ws_stats.insert_chart(chart_offset_y, 1, chart_settore)
    ws_stats.insert_chart(chart_offset_y, 13, chart_azienda)

    row_pie_tables = chart_offset_y + 26
    
    fmt_tbl_hdr = workbook.add_format({'bold': True, 'font_color': 'white', 'fg_color': '#4F81BD', 'border': 1, 'align': 'center'})
    fmt_tbl_cell = workbook.add_format({'num_format': '0.00%', 'border': 1, 'align': 'right'})
    fmt_tbl_lbl = workbook.add_format({'border': 1, 'align': 'left', 'size': 9})

    col_tbl_set = 9   
    col_tbl_az = 21   

    # 🟢 ALLARGA LE COLONNE TESTUALI DELLE TABELLE AFFIANCATE (Ora si legge tutto!)
    ws_stats.set_column(col_tbl_set, col_tbl_set, 28)
    ws_stats.set_column(col_tbl_az, col_tbl_az, 28)

    ws_stats.write(row_pie_tables, col_tbl_set, "Componente (Settore)", fmt_tbl_hdr)
    ws_stats.write(row_pie_tables, col_tbl_set + 1, "Quota 2024", fmt_tbl_hdr)
    ws_stats.write(row_pie_tables, col_tbl_az, "Componente (Azienda)", fmt_tbl_hdr)
    ws_stats.write(row_pie_tables, col_tbl_az + 1, "Quota 2024", fmt_tbl_hdr)

    for i, comp in enumerate(componenti_nomi.keys()):
        pos = posizioni_grafici[comp]
        ws_stats.write(row_pie_tables + 1 + i, col_tbl_set, comp, fmt_tbl_lbl)
        ws_stats.write_formula(row_pie_tables + 1 + i, col_tbl_set + 1, f"='7b. Stat. Indici Composizione'!{xlsxwriter.utility.xl_rowcol_to_cell(pos['settore_row'], len(all_years))}", fmt_tbl_cell)
        
        ws_stats.write(row_pie_tables + 1 + i, col_tbl_az, comp, fmt_tbl_lbl)
        ws_stats.write_formula(row_pie_tables + 1 + i, col_tbl_az + 1, f"='7b. Stat. Indici Composizione'!{xlsxwriter.utility.xl_rowcol_to_cell(pos['azienda_row'], len(all_years))}", fmt_tbl_cell)

    chart_pie_settore = workbook.add_chart({'type': 'pie'})
    chart_pie_settore.set_title({'name': 'Composizione 2024 - MEDIANE SETTORE'})
    chart_pie_settore.set_size({'width': 420, 'height': 340})
    chart_pie_settore.set_legend({'position': 'bottom', 'font': font_assi})

    chart_pie_target = workbook.add_chart({'type': 'pie'})
    chart_pie_target.set_title({'name': f'Composizione 2024 - {azienda_target}'})
    chart_pie_target.set_size({'width': 420, 'height': 340})
    chart_pie_target.set_legend({'position': 'bottom', 'font': font_assi})

    chart_pie_settore.add_series({
        'categories': ['7b. Stat. Indici Composizione', row_pie_tables + 1, col_tbl_set, row_pie_tables + len(componenti_nomi), col_tbl_set],
        'values': ['7b. Stat. Indici Composizione', row_pie_tables + 1, col_tbl_set + 1, row_pie_tables + len(componenti_nomi), col_tbl_set + 1]
    })

    chart_pie_target.add_series({
        'categories': ['7b. Stat. Indici Composizione', row_pie_tables + 1, col_tbl_az, row_pie_tables + len(componenti_nomi), col_tbl_az],
        'values': ['7b. Stat. Indici Composizione', row_pie_tables + 1, col_tbl_az + 1, row_pie_tables + len(componenti_nomi), col_tbl_az + 1]
    })

    ws_stats.insert_chart(row_pie_tables, 1, chart_pie_settore)
    ws_stats.insert_chart(row_pie_tables, 13, chart_pie_target)

    ws_stats.ignore_errors({'formula_differs': 'A1:Z500', 'number_stored_as_text': 'A1:Z500'})
    
    writer.close()
    output_buffer.seek(0)
    return output_buffer



def elabora_capitolo_7(df_filtered, azienda_target):
    import io
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import zipfile

    # --- FUNZIONI DI SUPPORTO ---

    # *** MODIFICA RISPETTO ALL'ORIGINALE ***
    # Il punteggio ora rispecchia il rango del terzile:
    #   1 = primo terzile  = valori più alti (MIGLIORI) per metriche dirette
    #   2 = secondo terzile = valori medi
    #   3 = terzo terzile  = valori più bassi (PEGGIORI)
    # Nell'originale era invertito: 3 = migliore, 1 = peggiore.
    # Il default NaN diventa 3 (peggiore) invece di 1.
    def calcola_punteggi_diretto(val, t1, t2):
        if pd.isna(val): return 3  # MODIFICATO: era return 1
        if val >= t2: return 1     # MODIFICATO: era return 3
        elif val >= t1: return 2
        else: return 3             # MODIFICATO: era return 1

    def calcola_punteggi_inverso(val, t1, t2):
        if pd.isna(val): return 3  # MODIFICATO: era return 1
        if val <= t1: return 1     # MODIFICATO: era return 3  (gearing basso = 1° terzile = migliore)
        elif val <= t2: return 2
        else: return 3             # MODIFICATO: era return 1

    # assegna_lettera INVARIATA: usata per il Benchmark Totale (Sum_Lettere 3-9, scala A=3/B=2/C=1)
    def assegna_lettera(punti, soglia_A=8, soglia_B=5):
        if pd.isna(punti): return 'C'
        if punti >= soglia_A: return 'A'
        elif punti >= soglia_B: return 'B'
        else: return 'C'

    # *** NUOVA FUNZIONE RISPETTO ALL'ORIGINALE ***
    # Usata per le 3 aree (Eco, Fin, Pat) con il nuovo sistema punteggio 1=migliore:
    #   Sum area da 3 (tre 1° terzili = eccellenza) a 9 (tre 3° terzili = vulnerabilità)
    #   A ≤ 4 | B ≤ 7 | C > 7 — produce gli stessi A/B/C dell'originale
    def assegna_lettera_area(punti):
        if pd.isna(punti): return 'C'
        return 'A' if punti <= 4 else ('B' if punti <= 7 else 'C')

    # punti_da_lettera INVARIATO: converte le lettere di area in punti per il Benchmark Totale
    def punti_da_lettera(lettera):
        if lettera == 'A': return 3
        elif lettera == 'B': return 2
        else: return 1

    df_raw = df_filtered.copy()

    def trova_col(keywords, exclude=None):
        if exclude is None: exclude = []
        for c in df_raw.columns:
            c_lower = str(c).lower()
            if all(k.lower() in c_lower for k in keywords) and not any(e.lower() in c_lower for e in exclude):
                return c
        return None

    col_regione = trova_col(['nuts', '2'])
    if not col_regione: col_regione = trova_col(['nuts'])
    if not col_regione:
        df_raw['Regione_Fallback'] = "N.D."
        col_regione = 'Regione_Fallback'

    col_societa = 'Società ID'
    df_raw.insert(1, col_societa, range(1, len(df_raw) + 1))
    
    col_ragione_sociale = trova_col(['ragione'])

    # Ricreiamo le metriche usate per il ranking
    metriche_utili = {
        'M. Profitto 2024': trova_col(['marg', 'profitto', '2024']),
        'M. EBITDA 2024': trova_col(['marg', 'ebitda', '2024']),
        'M. EBIT 2024': trova_col(['marg', 'ebit', '2024'], exclude=['ebitda']),
        'Rotazione C.Inv. 2024': trova_col(['rotazione', '2024']),
        'Quick Ratio 2024': trova_col(['quick', '2024']),
        'Current Ratio 2024': trova_col(['current', '2024']),
        'Indice 1° Liv. 2024': trova_col(['struttura 1', '2024']),
        'Indice 2° Liv. 2024': trova_col(['struttura 2', '2024']),
        'Gearing 2024': trova_col(['gearing', '2024'])
    }

    df = pd.DataFrame()
    df[col_ragione_sociale] = df_raw[col_ragione_sociale]
    df[col_societa] = df_raw[col_societa]
    df[col_regione] = df_raw[col_regione]
    
    for nome_bello, col_brutta in metriche_utili.items():
        if col_brutta:
            df[nome_bello] = pd.to_numeric(df_raw[col_brutta], errors='coerce')
        else:
            df[nome_bello] = np.nan

    col_lettere = ['Benchmark Economico', 'Benchmark Finanziario', 'Benchmark Patrimoniale', 'Benchmark Totale', 'Rating Combinato']

    categorie_kpi = {
        'Equilibrio_Economico': {
            'Prof_Mg': ('M. Profitto 2024', False),
            'EBITDA_Mg': ('M. EBITDA 2024', False),
            'EBIT_Mg': ('M. EBIT 2024', False)
        },
        'Equilibrio_Finanziario': {
            'Rotazione_Cap': ('Rotazione C.Inv. 2024', False),
            'Quick_Rat': ('Quick Ratio 2024', False),
            'Current_Rat': ('Current Ratio 2024', False)
        },
        'Equilibrio_Patrimoniale': {
            'IndStrut1': ('Indice 1° Liv. 2024', False),
            'IndStrut2': ('Indice 2° Liv. 2024', False),
            'Gearing': ('Gearing 2024', True) 
        }
    }

    tutti_kpi_cols = []
    for kpi_dict in categorie_kpi.values():
        tutti_kpi_cols.extend([v[0] for v in kpi_dict.values()])

    metriche_inverse = ['Gearing 2024']

    for m in tutti_kpi_cols:
        if m in df.columns:
            t1 = df[m].quantile(1/3)
            t2 = df[m].quantile(2/3)
            if m in metriche_inverse:
                df[f'Pts_{m}'] = df[m].apply(lambda x: calcola_punteggi_inverso(x, t1, t2))
            else:
                df[f'Pts_{m}'] = df[m].apply(lambda x: calcola_punteggi_diretto(x, t1, t2))

    if all(f'Pts_{x}' in df.columns for x in ['M. Profitto 2024', 'M. EBITDA 2024', 'M. EBIT 2024']):
        df['Sum_Eco'] = df['Pts_M. Profitto 2024'] + df['Pts_M. EBITDA 2024'] + df['Pts_M. EBIT 2024']
        df['Benchmark Economico'] = df['Sum_Eco'].apply(assegna_lettera_area)  # MODIFICATO: era assegna_lettera(x, 8, 5)
    else: df['Benchmark Economico'] = 'C'

    if all(f'Pts_{x}' in df.columns for x in ['Rotazione C.Inv. 2024', 'Quick Ratio 2024', 'Current Ratio 2024']):
        df['Sum_Fin'] = df['Pts_Rotazione C.Inv. 2024'] + df['Pts_Quick Ratio 2024'] + df['Pts_Current Ratio 2024']
        df['Benchmark Finanziario'] = df['Sum_Fin'].apply(assegna_lettera_area)  # MODIFICATO: era assegna_lettera(x, 8, 5)
    else: df['Benchmark Finanziario'] = 'C'

    if all(f'Pts_{x}' in df.columns for x in ['Indice 1° Liv. 2024', 'Indice 2° Liv. 2024', 'Gearing 2024']):
        df['Sum_Pat'] = df['Pts_Indice 1° Liv. 2024'] + df['Pts_Indice 2° Liv. 2024'] + df['Pts_Gearing 2024']
        df['Benchmark Patrimoniale'] = df['Sum_Pat'].apply(assegna_lettera_area)  # MODIFICATO: era assegna_lettera(x, 8, 5)
    else: df['Benchmark Patrimoniale'] = 'C'

    df['Sum_Lettere'] = df['Benchmark Economico'].apply(punti_da_lettera) + \
                        df['Benchmark Finanziario'].apply(punti_da_lettera) + \
                        df['Benchmark Patrimoniale'].apply(punti_da_lettera)

    df['Benchmark Totale'] = df['Sum_Lettere'].apply(lambda x: assegna_lettera(x, 8, 5))
    # 🟢 FIX: Ordine istituzionale delle stringhe (Eco + Patr + Fin)
    df['Rating Combinato'] = df['Benchmark Economico'] + df['Benchmark Patrimoniale'] + df['Benchmark Finanziario']

    df_master = df.dropna(subset=[col_ragione_sociale]).copy()
    colonne_da_tenere = [col_ragione_sociale, col_societa, col_regione] + [c for c in col_lettere if c in df.columns] + tutti_kpi_cols

    def calcola_rank(dataframe, col_kpi, base_name, lower_is_better):
        asc_order = True if lower_is_better else False
        dataframe[f'RANK_{base_name}_NAZ'] = dataframe[col_kpi].rank(ascending=asc_order, method='min')
        dataframe[f'RANK_{base_name}_REG'] = dataframe.groupby(col_regione, observed=False)[col_kpi].rank(ascending=asc_order, method='min')

    # *** NUOVO: pre-calcolo target e sottoinsieme regionale per le mediane regionali ***
    df_target_row = df_master[df_master[col_ragione_sociale].astype(str).str.lower().str.contains(azienda_target.lower().strip(), na=False)]
    regione_target = df_target_row.iloc[0][col_regione] if not df_target_row.empty else None
    df_regione_master = df_master[df_master[col_regione] == regione_target] if regione_target is not None else pd.DataFrame()
    nome_reg_pulita = str(regione_target).split(' - ')[-1] if regione_target else 'Regione N.D.'

    # Creazione del buffer ZIP per ospitare i 3 file Excel
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for categoria, kpi_dict in categorie_kpi.items():
            file_name = f'7_Ranking_Aziendale_{categoria}.xlsx'
            df_cat = df_master.copy()
            
            for kpi_name, (kpi_col, is_lower_better) in kpi_dict.items():
                if kpi_col in df_cat.columns:
                    calcola_rank(df_cat, kpi_col, kpi_name, is_lower_better)
            
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                if 'Rating Combinato' in df_cat.columns:
                    # ORA TENIAMO SIA I RANK NAZIONALI CHE QUELLI REGIONALI
                    colonne_rank = [c for c in df_cat.columns if '_NAZ' in c or '_REG' in c]
                    # Per lo spareggio nel foglio TOP RATING usiamo il primo rank nazionale
                    colonne_rank_naz = [c for c in df_cat.columns if '_NAZ' in c]
                    col_rank_spareggio = colonne_rank_naz[0] if colonne_rank_naz else col_societa
                    
                    df_rating = df_cat.sort_values(
                        by=['Rating Combinato', 'Benchmark Totale', col_rank_spareggio], 
                        ascending=[True, True, True] 
                    )
                    cols = [col_ragione_sociale] + [c for c in df_rating.columns if c in colonne_da_tenere + colonne_rank and c != col_ragione_sociale]
                    df_rating[cols].to_excel(writer, sheet_name='TOP_RATING_ABC', index=False)
                
                for kpi_name, (kpi_col, _) in kpi_dict.items():
                    if kpi_col in df_cat.columns:
                        df_naz = df_cat.sort_values(by=f'RANK_{kpi_name}_NAZ').dropna(subset=[kpi_col])
                        cols_naz = [col_ragione_sociale] + [c for c in df_naz.columns if c in colonne_da_tenere + colonne_rank and c != col_ragione_sociale]
                        df_naz[cols_naz].to_excel(writer, sheet_name=f'{kpi_name}_Nazionale', index=False)
                        
                        df_reg = df_cat.sort_values(by=[col_regione, f'RANK_{kpi_name}_REG']).dropna(subset=[kpi_col])
                        cols_reg = [col_ragione_sociale] + [c for c in df_reg.columns if c in colonne_da_tenere + colonne_rank and c != col_ragione_sociale]
                        df_reg[cols_reg].to_excel(writer, sheet_name=f'{kpi_name}_Regionale', index=False)

            excel_buffer.seek(0)
            wb = openpyxl.load_workbook(excel_buffer)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal='center')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # *** NUOVO: foglio "Riepilogo_Benchmark" con mediane nazionali e regionali ***
            # Specchio del confronto Italia / Regione / Azienda che report_breve_corp usa nei grafici e tabelle.
            ws_bench = wb.create_sheet("Riepilogo_Benchmark", 0)
            headers_bench = [
                'Metrica',
                'Mediana Nazionale',
                f'Mediana {nome_reg_pulita}',
                f'Valore {azienda_target[:25]}',
                'Rank Naz.',
                'Rank Reg.'
            ]
            for ci, h in enumerate(headers_bench, 1):
                ws_bench.cell(row=1, column=ci, value=h)  # header styling applicato dal loop sotto

            for ri, (kpi_name, (kpi_col, is_lower)) in enumerate(kpi_dict.items(), 2):
                if kpi_col not in df_master.columns:
                    continue
                med_naz = df_master[kpi_col].median()
                med_reg = df_regione_master[kpi_col].median() if not df_regione_master.empty else np.nan
                val_az  = df_target_row.iloc[0][kpi_col] if not df_target_row.empty else np.nan

                # rank nazionale
                tot_naz = int(df_master[kpi_col].notna().sum())
                if not df_target_row.empty:
                    _rn_series = df_master[kpi_col].rank(ascending=is_lower, method='min')
                    _idx = df_target_row.index[0]
                    _rn = _rn_series.get(_idx, np.nan)
                    rank_naz_str = f"{int(_rn)}/{tot_naz}" if pd.notna(_rn) else "n.d."
                else:
                    rank_naz_str = "n.d."

                # rank regionale
                if not df_regione_master.empty and not df_target_row.empty:
                    tot_reg = int(df_regione_master[kpi_col].notna().sum())
                    _rr_series = df_regione_master[kpi_col].rank(ascending=is_lower, method='min')
                    _rr = _rr_series.get(df_target_row.index[0], np.nan)
                    rank_reg_str = f"{int(_rr)}/{tot_reg}" if pd.notna(_rr) else "n.d."
                else:
                    rank_reg_str = "n.d."

                row_vals = [
                    kpi_name,
                    round(float(med_naz), 4) if pd.notna(med_naz) else None,
                    round(float(med_reg), 4) if pd.notna(med_reg) else None,
                    round(float(val_az),  4) if pd.notna(val_az)  else None,
                    rank_naz_str,
                    rank_reg_str,
                ]
                for ci, val in enumerate(row_vals, 1):
                    cell = ws_bench.cell(row=ri, column=ci, value=val)
                    cell.border = thin_border
                    if ci in [2, 3, 4] and val is not None:
                        cell.number_format = '#,##0.0000'
                    cell.alignment = center_align if ci > 1 else Alignment(horizontal='left')

            for col in ws_bench.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=12)
                ws_bench.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = center_align
                    cell.border = thin_border

                # --- AUTOMATISMO: EVIDENZIAZIONE AZIENDA TARGET NEL RANKING ---
                fill_target_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                font_target_bold = Font(bold=True)

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    # La prima colonna (A) contiene sempre la Ragione Sociale dell'azienda
                    if row[0].value and azienda_target.lower().strip() in str(row[0].value).lower().strip():
                        for cell in row:
                            cell.fill = fill_target_yellow
                            cell.font = font_target_bold

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    
                    if col[0].value and ("RANK" in str(col[0].value) or "Rating" in str(col[0].value) or "Benchmark" in str(col[0].value)):
                        for cell in col:
                            cell.alignment = center_align
                    
                    for cell in col:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 3, 50)

            final_excel_buffer = io.BytesIO()
            wb.save(final_excel_buffer)
            final_excel_buffer.seek(0)
            
            zip_file.writestr(file_name, final_excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer

# ==========================================
# 2. INTERFACCIA WEB (Il "Capitolo 0" + Filtri)
# ==========================================

uploaded_file = st.file_uploader("Trascina qui l'export grezzo (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    with st.spinner("Verifica formato, lettura e pulizia dati in corso..."):
            
        # --- 1. VERIFICA PRESENZA DEL FOGLIO CORRETTO ---
        try:
            xls = pd.ExcelFile(uploaded_file, engine='calamine')
        except Exception as e:
            st.error("❌ ERRORE: Impossibile leggere il file Excel.")
            st.warning(f"🛠️ DETTAGLIO TECNICO: {e}")
            st.stop()
            
        fogli_disponibili = xls.sheet_names
        target_sheet = None
        if "Risultati" in fogli_disponibili:
            target_sheet = "Risultati"
        elif "Results" in fogli_disponibili:
            target_sheet = "Results"
            
        if target_sheet is None:
            st.error("❌ ERRORE: Foglio 'Risultati' o 'Results' non trovato. ESPORTARE DA ORBIS CON I FILTRI E LA LISTA UNIVERSAL.")
            st.stop()

        # ==========================================
        # 🟢 NUOVO AUTOMATISMO: ESTRAZIONE SETTORE E FIX ACCENTI
        # ==========================================
        nomi_foglio_strategia = ['search strategy', 'search dataset', 'sommario ricerca', 'sommario', 'strategia ricerca']
        foglio_strategia_trovato = [s for s in fogli_disponibili if s.lower().strip() in nomi_foglio_strategia or any(k in s.lower() for k in ['strategy', 'sommario'])]
        
        settore_estratto = "Settore Non Rilevato"
        
        if foglio_strategia_trovato:
            try:
                # Leggiamo il foglio riepilogativo per estrarre la riga NACE
                df_strat = pd.read_excel(xls, sheet_name=foglio_strategia_trovato[0], header=None)
                for idx, row in df_strat.iterrows():
                    riga_testo_unita = " ".join(row.dropna().astype(str)).lower()
                    if 'nace' in riga_testo_unita or 'codici primari' in riga_testo_unita or 'ateco' in riga_testo_unita:
                        valori_riga = row.dropna().astype(str).tolist() # <--- RIGA MODIFICATA
                        celle_pulite = [str(c).strip() for c in valori_riga if '-' in str(c) and len(str(c)) > 10] # <--- RIGA MODIFICATA
                        if celle_pulite:
                            settore_estratto = celle_pulite[0]
                            st.session_state['universo_orbis'] = valori_riga[-1].strip() # <--- RIGA AGGIUNTA
                            break
                        else:
                            celle_lunghe = [str(c).strip() for c in row.dropna() if len(str(c)) > 15]
                            if len(celle_lunghe) >= 2:
                                settore_estratto = celle_lunghe[1]
                                break
            except Exception:
                settore_estratto = "Errore durante la lettura del sommario"
        
        # Correzione accenti (es. Forlì, Società) sul testo appena estratto
        fix_accenti = {'Ã¬': 'ì', 'Ã¨': 'è', 'Ã©': 'é', 'Ã²': 'ò', 'Ã¹': 'ù', 'Ã ': 'à', 'Ã': 'à'}
        for rotto, giusto in fix_accenti.items():
            settore_estratto = settore_estratto.replace(rotto, giusto)
            
        st.session_state['settore_estratto'] = settore_estratto

        # Lettura del foglio corretto (i risultati veri e propri)
        df_orbis = pd.read_excel(xls, sheet_name=target_sheet)


        # --- 2. VERIFICA STRUTTURA DELLE COLONNE ---
        # La tua "Lista Universal" rigorosa AGGIORNATA
        colonne_attese = [
            'Ragione socialeCaratteri latini', 'Numero BvD ID', 'Forma giuridica nazionale', 
            'NUTS1', 'NUTS2', 'NUTS3', 'Numero dipendenti 2024', 
            'Totale valore della produzione migl EUR 2024', 'Totale valore della produzione migl EUR 2023', 'Totale valore della produzione migl EUR 2022', 'Totale valore della produzione migl EUR 2021', 
            'Totale Attivo migl EUR 2024', 'Totale Attivo migl EUR 2023', 'Totale Attivo migl EUR 2022', 'Totale Attivo migl EUR 2021', 
            'Margine di Profitto (*) % 2024', 'Margine di Profitto (*) % 2023', 'Margine di Profitto (*) % 2022', 'Margine di Profitto (*) % 2021', 
            'Margine EBITDA (*) % 2024', 'Margine EBITDA (*) % 2023', 'Margine EBITDA (*) % 2022', 'Margine EBITDA (*) % 2021', 
            'Margine EBIT (*) % 2024', 'Margine EBIT (*) % 2023', 'Margine EBIT (*) % 2022', 'Margine EBIT (*) % 2021', 
            'Indice di Struttura 1° livello (*) 2024', 'Indice di Struttura 1° livello (*) 2023', 'Indice di Struttura 1° livello (*) 2022', 'Indice di Struttura 1° livello (*) 2021', 
            'Indice di Struttura 2° livello (*) 2024', 'Indice di Struttura 2° livello (*) 2023', 'Indice di Struttura 2° livello (*) 2022', 'Indice di Struttura 2° livello (*) 2021', 
            'Gearing (*) % 2024', 'Gearing (*) % 2023', 'Gearing (*) % 2022', 'Gearing (*) % 2021', 
            'Current Ratio (*) 2024', 'Current Ratio (*) 2023', 'Current Ratio (*) 2022', 'Current Ratio (*) 2021', 
            'Quick Ratio (*) 2024', 'Quick Ratio (*) 2023', 'Quick Ratio (*) 2022', 'Quick Ratio (*) 2021', 
            'Indice di Rotazione del Capitale Investito (*) 2024', 'Indice di Rotazione del Capitale Investito (*) 2023', 'Indice di Rotazione del Capitale Investito (*) 2022', 'Indice di Rotazione del Capitale Investito (*) 2021',
            # 🟢 NUOVE COLONNE AGGIUNTE INTEGRATE NELLA VALIDAZIONE
            'Costo del venduto migl EUR 2024', 'Costo del venduto migl EUR 2023', 'Costo del venduto migl EUR 2022', 'Costo del venduto migl EUR 2021',
            'Oneri diversi di gestione migl EUR 2024', 'Oneri diversi di gestione migl EUR 2023', 'Oneri diversi di gestione migl EUR 2022', 'Oneri diversi di gestione migl EUR 2021',
            'Proventi/oneri finanziari migl EUR 2024', 'Proventi/oneri finanziari migl EUR 2023', 'Proventi/oneri finanziari migl EUR 2022', 'Proventi/oneri finanziari migl EUR 2021',
            'Totale imposte migl EUR 2024', 'Totale imposte migl EUR 2023', 'Totale imposte migl EUR 2022', 'Totale imposte migl EUR 2021',
            'Utile/Perdita al netto delle imposte migl EUR 2024', 'Utile/Perdita al netto delle imposte migl EUR 2023', 'Utile/Perdita al netto delle imposte migl EUR 2022', 'Utile/Perdita al netto delle imposte migl EUR 2021',
            'Codice fiscale/Partita IVA', 'Indirizzo sito web', 'Indirizzo e-mail'
        ]
            
        # Pulisce gli spazi laterali dalle colonne per un controllo accurato
        colonne_file = [str(c).strip() for c in df_orbis.columns]
        colonne_mancanti = [col for col in colonne_attese if col not in colonne_file]
            
        if len(colonne_mancanti) > 0:
            st.error("❌ ERRORE: Colonne non valide o mancanti. ESPORTARE DA ORBIS CON I FILTRI E LA LISTA UNIVERSAL.")
            # Opzionale: un menu a tendina per far vedere esattamente cosa manca (utile per il debug)
            with st.expander("Mostra i dettagli dell'errore (Colonne Mancanti)"):
                st.write(colonne_mancanti)
            st.stop()  # IL PROGRAMMA SI FERMA QUI, I PULSANTI NON VERRANNO GENERATI

        # --- 3. PULIZIA DATI E FILTRAGGIO VALORI (n.d., Rotazione e Gearing) ---
        righe_iniziali = len(df_orbis)
        col_att_24 = 'Totale Attivo migl EUR 2024'
        col_ric_24 = 'Totale valore della produzione migl EUR 2024'
        col_rot_24 = 'Indice di Rotazione del Capitale Investito (*) 2024'
        
        col_g24 = 'Gearing (*) % 2024'
        col_g23 = 'Gearing (*) % 2023'
        col_g22 = 'Gearing (*) % 2022'
        col_g21 = 'Gearing (*) % 2021'

        # Forza la conversione a numero per tutte le metriche chiave
        for col in [col_att_24, col_ric_24, col_rot_24, col_g24, col_g23, col_g22, col_g21]:
            if col in df_orbis.columns:
                df_orbis[col] = pd.to_numeric(df_orbis[col], errors='coerce')

        # --- FILTRO 1: ROTAZIONE E DATI BASE ---
        df_orbis = df_orbis.dropna(subset=[col_att_24, col_ric_24, col_rot_24])
        df_orbis = df_orbis[df_orbis[col_rot_24] > 0]
        
        righe_post_rotazione = len(df_orbis)
        scartate_rotazione = righe_iniziali - righe_post_rotazione
        
        # --- FILTRO 2: GEARING ---
        if col_g24 in df_orbis.columns:
            # Elimina chi ha Gearing nullo o negativo nel 2024
            df_orbis = df_orbis[(df_orbis[col_g24].notna()) & (df_orbis[col_g24] > 0)]
            
            # Nasconde gli zeri degli anni passati trasformandoli in 'n.d.'
            for col_g in [col_g23, col_g22, col_g21]:
                if col_g in df_orbis.columns:
                    df_orbis[col_g] = df_orbis[col_g].replace(0, np.nan)
        
        righe_finali = len(df_orbis)
        scartate_gearing = righe_post_rotazione - righe_finali
        
        # Calcolo dei totali per la dashboard
        righe_scartate = righe_iniziali - righe_finali

        # --- NUOVO AUTOMATISMO: TRADUZIONE GEOGRAFICA ENG -> ITA E FIX ACCENTI ---
        def traduci_valori_territoriali(valore):
            if pd.isna(valore): return valore
            v_str = str(valore)
            
            # 🟢 FIX MOJIBAKE: Corregge i caratteri accentati rotti dagli export
            fix_accenti = {
                'Ã¬': 'ì',   # Forlì
                'Ã¨': 'è',   # (generico)
                'Ã©': 'é',   # (generico)
                'Ã²': 'ò',   # (generico)
                'Ã¹': 'ù',   # Cantù
                'Ã ': 'à',   # (generico)
                'Ã': 'à'     # Fallback se lo spazio è saltato
            }
            for rotto, giusto in fix_accenti.items():
                v_str = v_str.replace(rotto, giusto)

            # Dizionario delle corrispondenze Inglese -> Italiano usate da ORBIS
            dizionario_geo = {
                'North-East': 'Nord Est', 'Northeast': 'Nord Est', 'North East': 'Nord Est',
                'North-West': 'Nord Ovest', 'Northwest': 'Nord Ovest', 'North West': 'Nord Ovest',
                'Center': 'Centro', 'Centre': 'Centro', 'Central': 'Centro',
                'South': 'Sud', 'Islands': 'Isole', 'Insular Italy': 'Isole', 'Insular': 'Isole',
                'South and Islands': 'Sud e Isole', 'South and Insular Italy': 'Sud e Isole',
                'Lombardy': 'Lombardia', 'Sicily': 'Sicilia', 'Sardinia': 'Sardegna',
                'Apulia': 'Puglia', 'Tuscany': 'Toscana', 'Piedmont': 'Piemonte'
            }
            for eng, ita in dizionario_geo.items():
                v_str = re.sub(re.escape(eng), ita, v_str, flags=re.IGNORECASE)
                
            return v_str
        
        # Applica la traduzione solo alle colonne che contengono dati geografici
        for colonna in df_orbis.columns:
            if any(k in str(colonna).lower() for k in ['nuts', 'regione', 'provincia', 'territor']):
                df_orbis[colonna] = df_orbis[colonna].apply(traduci_valori_territoriali)
            
        # --- 4. RENUMERAZIONE PROGRESSIVA DELLA PRIMA COLONNA ---
        # Resetta l'indice del dataframe eliminando i "buchi"
        df_orbis = df_orbis.reset_index(drop=True)
            
        # La primissima colonna di ORBIS (spesso chiamata "Unnamed: 0") contiene la numerazione
        prima_colonna = df_orbis.columns[0] 
        if "Ragione" not in str(prima_colonna): # Controllo di sicurezza
            # Rigenera la lista con i nuovi posti scalati (es. "1.", "2.", "3.")
            df_orbis[prima_colonna] = [f"{i}." for i in range(1, len(df_orbis) + 1)]

        righe_finali = len(df_orbis)
        righe_scartate = righe_iniziali - righe_finali


    # --- SELEZIONE AUTOMATICA E INTELLIGENTE DELL'AZIENDA TARGET ---
        col_ragione_sociale = [c for c in df_orbis.columns if 'ragione' in str(c).lower()][0]

        st.markdown("### 🎯 Impostazione Azienda Target")
        ricerca_manuale = st.text_input(
            "Vuoi analizzare un'azienda specifica? (Opzionale)", 
            placeholder="Es: scrivi la Ragione Sociale... Lascia vuoto per l'auto-selezione intelligente."
        )

        azienda_target = None

        # --- 1. TENTATIVO DI RICERCA MANUALE ---
        if ricerca_manuale.strip():
            df_match = df_orbis[df_orbis[col_ragione_sociale].astype(str).str.lower().str.contains(ricerca_manuale.lower().strip(), na=False)]
            
            if not df_match.empty:
                azienda_target = df_match.iloc[0][col_ragione_sociale]
                st.success(f"✅ **Azienda Target forzata manualmente:** {azienda_target}")
            else:
                st.warning(f"⚠️ Nessuna azienda trovata contenente '{ricerca_manuale}'. Procedo con la selezione automatica.")

        # --- 2. SELEZIONE AUTOMATICA (Se l'utente non ha scritto nulla o non l'ha trovata) ---
        if azienda_target is None:
            
            # ⚠️ DA QUI IN POI TUTTO IL TUO CODICE HA 4 SPAZI (1 TAB) DI INDENTAZIONE ⚠️

            # 1. Pesca TUTTI gli indicatori, i margini e i ratio del 2024 presenti nel file
            colonne_kpi_2024 = [c for c in df_orbis.columns if '2024' in str(c) and any(x in str(c).lower() for x in ['margine', 'indice', 'ratio', 'gearing'])]
            
            # Tieni solo chi ha i dati completi (se il file è vuoto fa un fallback)
            df_candidati = df_orbis.dropna(subset=colonne_kpi_2024).copy()
            if df_candidati.empty: 
                df_candidati = df_orbis.copy()
                
            df_candidati['Score_Anomalia_Totale'] = 0
            df_candidati['Picco_Anomalia_Singola'] = 0 
            
            # 2. Calcola lo scostamento (Z-Score) per ogni singola metrica
            for col in colonne_kpi_2024:
                df_candidati[col] = pd.to_numeric(df_candidati[col], errors='coerce')
                mediana_settore = df_candidati[col].median()
                deviazione_std = df_candidati[col].std()
                
                if pd.notna(deviazione_std) and deviazione_std > 0:
                    scostamento = abs(df_candidati[col] - mediana_settore) / deviazione_std
                    df_candidati['Score_Anomalia_Totale'] += scostamento
                    df_candidati['Picco_Anomalia_Singola'] = df_candidati[['Picco_Anomalia_Singola']].assign(new=scostamento).max(axis=1)

            # 3. FILTRO 1 (No Outlier): Elimina le aziende che hanno anche solo UN indicatore fuori di testa
            df_puliti = df_candidati[df_candidati['Picco_Anomalia_Singola'] <= 1.5].copy()
            
            if df_puliti.empty: 
                df_puliti = df_candidati.copy()

            # MINIMO INTERVENTO: Preferenza per chi ha dichiarato i dipendenti
            if 'Numero dipendenti 2024' in df_puliti.columns:
                df_con_dipendenti = df_puliti[pd.to_numeric(df_puliti['Numero dipendenti 2024'], errors='coerce').notna()]
                if not df_con_dipendenti.empty:
                    df_puliti = df_con_dipendenti.copy()

            # 4. FILTRO 2 (No Perfetti): Ordiniamo dal più "mediano" al più "strano". 
            df_puliti = df_puliti.sort_values('Score_Anomalia_Totale', ascending=True).reset_index(drop=True)
            
            indice_genuino = max(1, len(df_puliti) // 7) 
            
            azienda_target = df_puliti.iloc[indice_genuino][col_ragione_sociale]
            
            # Messaggio a schermo opzionale per farti vedere chi ha scelto
            st.info(f"🤖 **Azienda Target Auto-Selezionata:** {azienda_target} (Rappresentativa del settore)")

    # ==========================================
    # DASHBOARD DATI CARICATI
    # ==========================================
    st.markdown("### 📊 Analisi Qualità del Dato")
    
    # 🟢 VISUALIZZAZIONE INFO SETTORE (Recuperata dallo step precedente)
    settore_visualizzato = st.session_state.get('settore_estratto', 'Settore Non Rilevato')
    st.info(f"📋 **Filtro Settore Rilevato:** {settore_visualizzato}")
    
    # Creiamo 3 colonne visive
    col1, col2, col3 = st.columns(3)
        
    with col1:
        st.metric(label="📄 Aziende Iniziali", value=righe_iniziali)
            
    with col2:
    # Mostra in rosso i dati scartati, con il dettaglio sotto!
        st.metric(
            label="🗑️ Aziende Scartate", 
            value=righe_scartate, 
            delta=f"-{scartate_rotazione} Rotazione | -{scartate_gearing} Gearing", 
            delta_color="inverse"
        )
            
    with col3:
    # Mostra in verde i dati pronti per l'analisi
        st.metric(
            label="✅ Aziende Valide", 
            value=righe_finali, 
            delta="Pronte per l'analisi", 
            delta_color="normal"
        )
    st.divider()

    # ==========================================
    # 3. GESTIONE DELLE TABS (I Pulsanti)
    # ==========================================
    
    # Crea le schede per i vari capitoli + Tab per il mega download
    tab9, tab8, tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab7_5 = st.tabs([
        "📄 Genera Report + PPT",
        "⭐ Scarica Tutto",
        "Cap 1: Forma Giur.",
        "Cap 2: Territorio",
        "Cap 3: Economico",
        "Cap 4: Patrimoniale",
        "Cap 5: Finanziario",
        "Cap 6: Benchmark",
        "Cap 7: Ranking",
        "Cap 7.5: Composizione",
    ])

    # --- SCHEDA CAPITOLO 1 ---
    with tab1:
        st.subheader("1. Analisi Forma Giuridica")
        st.write("Genera l'analisi aggregata per S.p.A. e S.r.l.")
        if st.button("Genera Capitolo 1", type="primary", key="btn_cap1"):
            with st.spinner("Creazione tabelle in corso..."):
                excel_cap1 = elabora_capitolo_1(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '1. Forma Giuridica'",
                    data=excel_cap1,
                    file_name="1_Forma_Giuridica.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 2 ---
    with tab2:
        st.subheader("2. Analisi Territoriale")
        st.write("Genera il report aggregato per NUTS2 (Regioni e Macroregioni) con grafici a torta e istogrammi.")
        if st.button("Genera Capitolo 2", type="primary", key="btn_cap2"):
            with st.spinner("Creazione tabelle e grafici territoriali..."):
                excel_cap2 = elabora_capitolo_2(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '2. Ripartizione Territoriale'",
                    data=excel_cap2,
                    file_name="2_Ripartizione_Territoriale.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 3 ---
    with tab3:
        st.subheader("3. Equilibrio Economico & Svil. Dimensionale")
        st.write("Genera analisi approfondite su Margini (Profitto, EBITDA, EBIT) e Ricavi.")
        if st.button("Genera Capitolo 3", type="primary", key="btn_cap3"):
            with st.spinner("Calcolo indici economici..."):
                excel_cap3 = elabora_capitolo_3(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '3. Equilibrio Economico'",
                    data=excel_cap3,
                    file_name="3_Equilibrio_Economico.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 4 ---
    with tab4:
        st.subheader("4. Equilibrio Patrimoniale")
        st.write("Genera analisi sugli indici di Struttura e Gearing.")
        if st.button("Genera Capitolo 4", type="primary", key="btn_cap4"):
            with st.spinner("Calcolo metriche patrimoniali..."):
                excel_cap4 = elabora_capitolo_4(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '4. Equilibrio Patrimoniale'",
                    data=excel_cap4,
                    file_name="4_Equilibrio_Patrimoniale.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 5 ---
    with tab5:
        st.subheader("5. Equilibrio Finanziario")
        st.write("Genera analisi sugli indici di liquidità (Current, Quick) e rotazione.")
        if st.button("Genera Capitolo 5", type="primary", key="btn_cap5"):
            with st.spinner("Calcolo metriche finanziarie..."):
                excel_cap5 = elabora_capitolo_5(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '5. Equilibrio Finanziario'",
                    data=excel_cap5,
                    file_name="5_Equilibrio_Finanziario.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 6 ---
    with tab6:
        st.subheader("6. Rating e Benchmark")
        st.write("Genera il cruscotto finale con calcolo terzili, assegnazione rating A-B-C e pivot territoriali.")
        if st.button("Genera Capitolo 6", type="primary", key="btn_cap6"):
            with st.spinner("Calcolo benchmark e assegnazione rating..."):
                excel_cap6 = elabora_capitolo_6(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '6. Benchmark'",
                    data=excel_cap6,
                    file_name="6_Benchmark.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # --- SCHEDA CAPITOLO 7 ---
    with tab7:
        st.subheader("7. Ranking Aziendale")
        st.write("Elabora le classifiche nazionali e regionali. Scaricherai un archivio ZIP contenente 3 file Excel.")
        if st.button("Genera Capitolo 7 (Pack ZIP)", type="primary", key="btn_cap7"):
            with st.spinner("Creazione ranking e compressione file..."):
                zip_cap7 = elabora_capitolo_7(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '7. Pacchetto Ranking' (ZIP)",
                    data=zip_cap7,
                    file_name="7_Ranking_Aziendale_Pack.zip",
                    mime="application/zip"
                )
    
    # ==========================================
    # 🟢 NUOVO TAB 7: DASHBOARD INDICI DI COMPOSIZIONE
    # ==========================================
    with tab7_5:
        st.subheader("7.5 Indici di Composizione")
        st.write("Analisi della composizione percentuale delle singole voci di costo e di utile sul fatturato complessivo.")
        if st.button("Genera Capitolo 7.5", type="primary", key="btn_cap7_5"):
            with st.spinner("Calcolo indici e generazione torte di composizione..."):
                excel_cap7_5 = elabora_capitolo_7_5(df_orbis, azienda_target)
                st.download_button(
                    label="📥 Scarica '7.5 Indici di Composizione'",
                    data=excel_cap7_5,
                    file_name="7.5_Indici_Composizione.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    # --- SCHEDA SCARICA TUTTO ---
    with tab8:
        st.subheader("⭐ Master Export: Tutti i Capitoli")
        st.write("Con un solo clic, Python elaborerà tutti i capitoli e ti restituirà un unico archivio ZIP contenente l'intero progetto.")
        if st.button("🚀 GENERA INTERO PROGETTO", type="primary", use_container_width=True, key="btn_all"):
            with st.spinner("Elaborazione massiva in corso... Mettiti comodo, potrebbe volerci qualche secondo!"):
                
                # Creiamo il mega ZIP in memoria
                master_zip_buffer = io.BytesIO()
                import zipfile # Assicurati sia importato
                
                with zipfile.ZipFile(master_zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as master_zip:
                    # 🟢 GENERAZIONE DEL FILE TXT CON IL NOME DEL SETTORE ESTRAZIONALE
                    testo_settore = st.session_state.get('settore_estratto', 'Settore Non Rilevato dall\'export')
                    contenuto_txt = f"PROGETTO FINHACK - REPORT GENERATO\n\nTarget Settore Industriale (Filtro NACE):\n{testo_settore}\n"
                    
                    # Scrive il file direttamente in memoria nell'archivio ZIP
                    master_zip.writestr("Info_Settore_Ricerca.txt", contenuto_txt)

                    # Eseguiamo e salviamo i primi 6 capitoli (il codice rimane invariato sotto)
                    master_zip.writestr("1_Forma_Giuridica.xlsx", elabora_capitolo_1(df_orbis, azienda_target).read())
                    master_zip.writestr("2_Ripartizione_Territoriale.xlsx", elabora_capitolo_2(df_orbis, azienda_target).read())
                    master_zip.writestr("3_Equilibrio_Economico.xlsx", elabora_capitolo_3(df_orbis, azienda_target).read())
                    master_zip.writestr("4_Equilibrio_Patrimoniale.xlsx", elabora_capitolo_4(df_orbis, azienda_target).read())
                    master_zip.writestr("5_Equilibrio_Finanziario.xlsx", elabora_capitolo_5(df_orbis, azienda_target).read())
                    master_zip.writestr("6_Benchmark.xlsx", elabora_capitolo_6(df_orbis, azienda_target).read())
                    master_zip.writestr("7.5_Indici_Composizione.xlsx", elabora_capitolo_7_5(df_orbis, azienda_target).read())
                    cap7_zip_buffer = elabora_capitolo_7(df_orbis, azienda_target)
                    with zipfile.ZipFile(cap7_zip_buffer, "r") as cap7_zip:
                        for nome_file in cap7_zip.namelist():
                            master_zip.writestr(nome_file, cap7_zip.read(nome_file))
                master_zip_buffer.seek(0)

                # =========================================================
                # ✂️ PULIZIA NOME FILE (Estrae solo i codici NACE numerici)
                # =========================================================
                # Cerca tutti i blocchi di 3 o 4 numeri seguiti da un trattino
                codici_file = re.findall(r'(\d{3,4})\s*-', str(testo_settore))
                
                if codici_file:
                    # Se trova codici, li unisce con un underscore (es. "6820_4511")
                    settore_pulito = "_".join(codici_file)
                else:
                    # Se non ne trova o il settore è "N.D.", pulisce i caratteri strani
                    settore_pulito = re.sub(r'[^a-zA-Z0-9]', '_', str(testo_settore)[:15]).strip('_')
                
                st.success("Tutti i capitoli elaborati con successo!")
                st.download_button(
                    label="📥 SCARICA PROGETTO COMPLETO (.zip)",
                    data=master_zip_buffer,
                    file_name=f"{settore_pulito}_FH_Analisi_Completa.zip",
                    mime="application/zip",
                    use_container_width=True
                )

    # --- SCHEDA GENERATORI DOCUMENTI (WORD E PPTX AFFIANCATI) ---
    with tab9:
        st.subheader("📄 Generazione Documenti Finali")
        st.markdown("Esporta i risultati in un report descrittivo completo o in una presentazione di sintesi.")
        
        # Crea due colonne di uguale larghezza
        col_word, col_ppt = st.columns(2)
        
        # ==========================================
        # 📝 COLONNA SINISTRA: REPORT WORD
        # ==========================================
        with col_word:
            st.info("Report testuale approfondito con tutte le analisi e le narrative.")

            attiva_watermark = st.toggle("🔒 Applica Watermark", value=False, help="Copre i dati sensibili del 2024 con logo F&V")

            if st.button("✨ GENERA REPORT WORD", type="primary", use_container_width=True, key="btn_word"):
                import os
                import io
                import zipfile
                import docx
                import jinja2
                import re
                
                percorso_template = "template_corp_doc_master.docx"
                if not os.path.exists(percorso_template):
                    st.error(f"❌ Impossibile trovare il file '{percorso_template}'.")
                else:
                    with st.spinner("Generazione Word in corso..."):
                        try:
                            virtual_zip_buffer = io.BytesIO()
                            with zipfile.ZipFile(virtual_zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as virtual_zip:
                                virtual_zip.writestr("1_Forma_Giuridica.xlsx", elabora_capitolo_1(df_orbis, azienda_target).read())
                                virtual_zip.writestr("2_Ripartizione_Territoriale.xlsx", elabora_capitolo_2(df_orbis, azienda_target).read())
                                virtual_zip.writestr("3_Equilibrio_Economico.xlsx", elabora_capitolo_3(df_orbis, azienda_target).read())
                                virtual_zip.writestr("4_Equilibrio_Patrimoniale.xlsx", elabora_capitolo_4(df_orbis, azienda_target).read())
                                virtual_zip.writestr("5_Equilibrio_Finanziario.xlsx", elabora_capitolo_5(df_orbis, azienda_target).read())
                                virtual_zip.writestr("6_Benchmark.xlsx", elabora_capitolo_6(df_orbis, azienda_target).read())
                            virtual_zip_buffer.seek(0)
                            
                            from report_corp import genera_report_word
                            settore = st.session_state.get('settore_estratto', 'N.D.')
                            universo = st.session_state.get('universo_orbis', 'N/D')
                            
                            word_finito = genera_report_word(virtual_zip_buffer, percorso_template, azienda_target, df_orbis, settore, universo, modalita_teaser=attiva_watermark)
                            
                            # PULIZIA NOME FILE
                            codici_file = re.findall(r'(\d{3,4})\s*-', str(settore))
                            if codici_file:
                                settore_pulito = "_".join(codici_file)
                            else:
                                settore_pulito = re.sub(r'[^a-zA-Z0-9]', '_', str(settore)[:15]).strip('_')
                            
                            suffisso_watermark = "_Watermark" if attiva_watermark else ""

                            st.success("Report generato con successo! 🎉")
                            st.download_button(
                                label="📥 SCARICA REPORT FINALE (.docx)",
                                data=word_finito,
                                # Inserito il suffisso dinamico prima dell'estensione
                                file_name=f"{settore_pulito}_Report_Base_{azienda_target}{suffisso_watermark}.docx",
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True,
                                key="dw_word"
                            )
                            
                        # --- L'ESTRATTORE DI CRASH ---
                        except jinja2.exceptions.TemplateSyntaxError as e:
                            st.error("❌ BASTA ANDARE ALLA CIECA. ABBIAMO TROVATO IL PUNTO ESATTO!")
                            try:
                                if hasattr(e, 'source') and e.source:
                                    linee = e.source.splitlines()
                                    riga = e.lineno - 1
                                    inizio = max(0, riga - 2)
                                    fine = min(len(linee), riga + 3)
                                    blocco_xml = " ".join(linee[inizio:fine])
                                    testo_puro = re.sub(r'<[^>]+>', ' ', blocco_xml)
                                    testo_puro = re.sub(r'\s+', ' ', testo_puro).strip()
                                    
                                    st.warning(f"📍 **L'ERRORE SI TROVA ESATTAMENTE IN MEZZO A QUESTE PAROLE:**")
                                    st.info(f"👉 ... {testo_puro} ...")
                                    st.markdown(f"**Dettaglio del motore:** `{e.message}`")
                                    st.markdown("💡 **SOLUZIONE DEFINITIVA:** Copia due o tre parole dal riquadro azzurro qui sopra. Vai nel tuo Word originale, fai `Ctrl+F` e cercale. Quello è il punto esatto in cui c'è una parentesi di troppo.")
                                else:
                                    st.error(f"Dettaglio Tecnico: {e.message}")
                            except Exception as fallback:
                                st.error(f"Errore tecnico: {e.message}")
                        except jinja2.exceptions.UndefinedError as e:
                            st.error("❌ VARIABILE SCONOSCIUTA!")
                            st.warning(f"Hai usato un tag che non esiste: **{e.message}**")
                        except Exception as e:
                            st.error(f"⚠️ Errore generico in fase di impaginazione: {str(e)}")

        # ==========================================
        # 📊 COLONNA DESTRA: PRESENTAZIONE PPTX
        # ==========================================
        with col_ppt:
            st.info("Slide di sintesi con grafici pronte per la presentazione al cliente.")
            # 👇 RIGA INVISIBILE PER PAREGGIARE L'ALTEZZA DEL TOGGLE A SINISTRA
            st.markdown("<div style='height: 38px;'></div>", unsafe_allow_html=True)
            if st.button("✨ GENERA PRESENTAZIONE PPTX", type="primary", use_container_width=True, key="btn_ppt"):
                import os
                import re
                
                percorso_template_ppt = "template_corp_ppt_master.pptx"
                if not os.path.exists(percorso_template_ppt):
                    st.error(f"❌ Impossibile trovare il file '{percorso_template_ppt}'.")
                else:
                    with st.spinner("Generazione slide in corso..."):
                        try:
                            # IMPORT MANCANTE INSERITO QUI:
                            from report_breve_corp import genera_presentazione_ppt
                            
                            settore = st.session_state.get('settore_estratto', 'N.D.')
                            universo = st.session_state.get('universo_orbis', 'N/D')
                            
                            # Calcolo del nome pulito per il file in download
                            codici_file = re.findall(r'(\d{3,4})\s*-', str(settore))
                            if codici_file:
                                settore_pulito = "_".join(codici_file)
                            else:
                                settore_pulito = re.sub(r'[^a-zA-Z0-9]', '_', str(settore)[:15]).strip('_')
                            
                            # Esecuzione script PowerPoint
                            ppt_finito = genera_presentazione_ppt(percorso_template_ppt, azienda_target, df_orbis, settore, universo)
                            
                            st.success("Presentazione generata con successo! 🎉")
                            st.download_button(
                                label="📥 SCARICA PRESENTAZIONE (.pptx)",
                                data=ppt_finito,
                                file_name=f"{settore_pulito}_Presentazione_{azienda_target}.pptx",
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                use_container_width=True,
                                key="dw_ppt"
                            )
                        except Exception as e:
                            st.error(f"⚠️ Errore durante la generazione del PPTX: {str(e)}")
